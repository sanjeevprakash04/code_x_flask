import pandas as pd
# import xlsxwriter
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import colors, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def IOChecklist(df):

    identifiers = ["FRQ", "MTR", "AIN", "AOUT", "DIN", "PID", "ALM", "VLV"]

    checklist_list = list()
    for i in range(len(df)):
        # gen_con = None
        identifier = df.loc[i, "Generator Function"]
        sort = str(df.loc[i, "Sort"])
        if identifier in identifiers:
                checklist_dict = {
                            'Sort'                      : str(df.loc[i, "Sort"]),
                            'Tag Part 1'                : df.loc[i, "TAG Part 2 (# Item)"],
                            'Tag Part 2'                : df.loc[i, "TAG Part 3 (Function code)"],
                            'Tag Part 3'                : df.loc[i, "TAG Part 4 (# Function serial)"],
                            'Tag'                       : f'{df.loc[i, "TAG Part 2 (# Item)"]}_{df.loc[i, "TAG Part 3 (Function code)"]}_{df.loc[i, "TAG Part 4 (# Function serial)"]}',
                            'Description'               : f'{df.loc[i, "Component Description"]} - {df.loc[i, "Process Description"]}',
                            'Main Power [KW]'           : df.loc[i, "Main Power [kW]"],
                            'Main Current [A]'          : df.loc[i, "Main Current [A]"],
                            'Main Voltage [V]'          : df.loc[i, "Main Voltage [V]"],
                            'Poles'                     : df.loc[i, "Poles"],
                            'Cos φ'                     : df.loc[i, "Cos φ (Efficiency)"],
                            'RPM'                       : df.loc[i, "RPM"],
                            'EU'                        : df.loc[i, "EU"],
                            'Min EU'                    : df.loc[i, "Min EU"],
                            'Max EU'                    : df.loc[i, "Max EU"],
                            'Raw Min'                   : df.loc[i, "Raw Min"],
                            'Raw Max'                   : df.loc[i, "Raw Max"],
                            'Signal Type'               : df.loc[i, "Signal Type"],
                            'PLC Address'               : df.loc[i, "PLC Address"],
                            'T. Func'                   : df.loc[i, "T. Funk"],
                            'Check CPU'                 : '',
                            'Check HMI'                 : ''

                            }
                checklist_list.append(checklist_dict)

        if df.loc[i, "Signal function NO/NC"] != pd.isna and sort == "3" and identifier not in identifiers:
            checklist_dict = {
                            'Sort'                      : str(df.loc[i, "Sort"]),
                            'Tag Part 1'                : df.loc[i, "TAG Part 2 (# Item)"],
                            'Tag Part 2'                : df.loc[i, "TAG Part 3 (Function code)"],
                            'Tag Part 3'                : df.loc[i, "TAG Part 4 (# Function serial)"],
                            'Tag'                       : f'{df.loc[i, "TAG Part 2 (# Item)"]}_{df.loc[i, "TAG Part 3 (Function code)"]}_{df.loc[i, "TAG Part 4 (# Function serial)"]}',
                            'Description'               : f'{df.loc[i, "Component Description"]} - {df.loc[i, "Process Description"]}',
                            'Main Power [KW]'           : df.loc[i, "Main Power [kW]"],
                            'Main Current [A]'          : df.loc[i, "Main Current [A]"],
                            'Main Voltage [V]'          : df.loc[i, "Main Voltage [V]"],
                            'Poles'                     : df.loc[i, "Poles"],
                            'Cos φ'                     : df.loc[i, "Cos φ (Efficiency)"],
                            'RPM'                       : df.loc[i, "RPM"],
                            'EU'                        : df.loc[i, "EU"],
                            'Min EU'                    : df.loc[i, "Min EU"],
                            'Max EU'                    : df.loc[i, "Max EU"],
                            'Raw Min'                   : df.loc[i, "Raw Min"],
                            'Raw Max'                   : df.loc[i, "Raw Max"],
                            'Signal Type'               : df.loc[i, "Signal Type"],
                            'PLC Address'               : df.loc[i, "PLC Address"],
                            'T. Func'                   : df.loc[i, "T. Funk"],
                            'Check CPU'                 : '',
                            'Check HMI'                 : ''

                            }
            checklist_list.append(checklist_dict)

        if sort == "1":
                checklist_dict = {
                            'Sort'                      : str(df.loc[i, "Sort"]),
                            'Tag Part 1'                : df.loc[i, "TAG Part 2 (# Item)"],
                            'Tag Part 2'                : '',
                            'Tag Part 3'                : '',
                            'Tag'                       : '',
                            'Description'               : df.loc[i, "Component Description"],
                            'Main Power [KW]'           : '',
                            'Main Current [A]'          : '',
                            'Main Voltage [V]'          : '',
                            'Poles'                     : '',
                            'Cos φ'                     : '',
                            'RPM'                       : '',
                            'EU'                        : '',
                            'Min EU'                    : '',
                            'Max EU'                    : '',
                            'Raw Min'                   : '',
                            'Raw Max'                   : '',
                            'Signal Type'               : '',
                            'PLC Address'               : '',
                            'T. Func'                   : '',
                            'Check CPU'                 : '',
                            'Check HMI'                 : ''

                            }
                
                checklist_list.append(checklist_dict)
    checklist_df = pd.DataFrame(checklist_list)
    checklist_df = checklist_df.replace(np.nan, '', regex=True)

    # def highlight_greaterthan(s, threshold, column):
    #     is_max = pd.Series(data=False, index=s.index)
    #     is_max[column] = s.loc[column] <= threshold
    #     return ['background-color: yellow' if is_max.any() else '' for v in is_max]


    # checklist_df.style.apply(highlight_greaterthan, threshold=1.0, column=['A'], axis=1)


    # def highlight_diff(tag):
    #     for i in range(len(checklist_df)):    
    #         tag = checklist_df.iloc[i, 0]
    #         value2 = checklist_df.iloc[i, 6]
    #         diff = 'background:yellow' if tag != value2 else 'background:white'
    #         return 'color: %s' % diff
    # checklist_df.style.applymap(highlight_diff)
    with pd.ExcelWriter(r'export/io-checklist/IO-Checklist.xlsx', engine='xlsxwriter') as writer:
        checklist_df.to_excel(writer, sheet_name='IO-Checklist', index=False)
        for column in checklist_df:
            column_width = max(checklist_df[column].astype(str).map(len).max(), len(column))
            col_idx = checklist_df.columns.get_loc(column)
            writer.sheets['IO-Checklist'].set_column(col_idx, col_idx, column_width)
    # writer = pd.ExcelWriter(r'export/io-checklist/IO-Checklist.xlsx', engine='xlsxwriter') # pylint: disable=abstract-class-instantiated

    # IO Checklist sheet
    # checklist_df.to_excel(writer, sheet_name='IO-Checklist', index=False)
    # for column in checklist_df:
    #     column_width = max(checklist_df[column].astype(str).map(len).max(), len(column))
    #     col_idx = checklist_df.columns.get_loc(column)
    #     writer.sheets['IO-Checklist'].set_column(col_idx, col_idx, column_width)

    # writer.save()


# ----------------Apply styles to sheets----------------
    filepath = r'export/io-checklist/IO-Checklist.xlsx'
    wb = load_workbook(filepath)
    ws = wb['IO-Checklist']

    #column_count = ws.max_column
    row_count = ws.max_row + 1

    # ----------------Styles for spreadsheet----------------
    bg_color = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    nrfill1 = PatternFill(fill_type='solid',
                    start_color='00FF8080',
                    end_color='00FF8080')
    fatfont = Font(name='Calibri Light',
                size=8,
                bold=True,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='FF000000')

    for x in range(row_count):
        if x % 2 != 0:
            for cell in ws[x:x]:
                cell.fill = bg_color

    #for idx, i in enumerate(ws[get_column_letter(sort+1)]):
    #    sort_nr = i.value
    #    if (sort_nr == "1") or (sort_nr == 1): # When Sort is 1
    #        for rows in ws.iter_rows(min_row=idx+1, max_row=idx+1, max_col=column_count):
    #            for cell in rows:
    #                cell.fill = nrfill1
    #                cell.font = fatfont
    
    wb.save(filepath)
    wb.close()