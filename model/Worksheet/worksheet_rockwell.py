import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import colors, Font, PatternFill, Border, Side

def getSheet(df, Cmd_SortRows, Cmd_CheckDups, Cmd_AssignCMNo, Cmd_CreateSheets):


    # Initialization of Dataframe   
    df.columns = df.iloc[2].astype(str) # Set the header to be at index 2
    #df = pd.DataFrame(df.iloc[3:, 1:]) # Remove the first 3 rows and the first column
    df = pd.DataFrame(df.iloc[3:]) # Remove the first 3 rows and the first column
    df = df.reset_index()

    # Creates lists for dataframes
    frq_list = list()
    mtr_list = list()
    vlv_list = list()
    din_list = list()
    ain_list = list()
    aout_list = list()
    pid_list = list()
    alm_list = list()
    fp_list = list()
    fp_vlv_list = list()
    aoi_safe_list = list()
    sum_list = list()
    ipAdr_list = list()

    # Status variable to display current program state
    global Sts_PrgStatus

    identifiers = ["FRQ", "MTR", "AIN", "AOUT", "DIN", "PID", "ALM", "VLV"]

    Sts_PrgStatus = "Please change Worksheet settings"

    def AssignCMNumber(): # ----------------Autogenerate Generator Config Number----------------
    ##### FRQ #####
        num_count = 0
        next_num = 0
        num_count_list = list()
        for i in range(len(df)): # Check is CM number is already assigned and get the highest value
            identifier = df.loc[i, "Generator Function"]
            if identifier == "FRQ" and pd.notna(df.loc[i, "CM No"]):
                num_count = df.loc[i, "CM No"]
                num_count_list.append(num_count)

        if num_count_list == []: # Assign starting CM number
            next_num = 1000
        else:
            next_num = max(num_count_list)

        for i in range(len(df)):
            identifier = str(df.loc[i, "Generator Function"])
            tag_part_3 = df.loc[i, "TAG Part 3 (Function code)"]
            if identifier == "FRQ" and pd.isna(df.loc[i, "CM No"]):
                next_num = next_num + 1
                df.loc[i, "CM No"] = str(next_num)
                for x in range(len(df)):
                    identifier = df.loc[x, "Generator Function"]
                    if (identifier == "FRQ_Ptc") and (df.loc[x, "TAG Part 3 (Function code)"] == tag_part_3):
                        df.loc[x, "CM No"] = str(next_num)
                    elif (identifier == "FRQ_Pro") and (df.loc[x, "TAG Part 3 (Function code)"] == tag_part_3):
                        df.loc[x, "CM No"] = str(next_num)
                    elif (identifier == "FRQ_R") and (df.loc[x, "TAG Part 3 (Function code)"] == tag_part_3):
                        df.loc[x, "CM No"] = str(next_num)
                    elif (identifier == "FRQ_Dis") and (df.loc[x, "TAG Part 3 (Function code)"] == tag_part_3):
                        df.loc[x, "CM No"] = str(next_num)
                    elif (identifier == "FRQ_Rot") and (df.loc[x, "TAG Part 3 (Function code)"] == tag_part_3):
                        df.loc[x, "CM No"] = str(next_num)
                    elif (identifier == "FRQ_ETH") and (df.loc[x, "TAG Part 3 (Function code)"] == tag_part_3):
                        df.loc[x, "CM No"] = str(next_num)

    ##### MTR #####
        num_count = 0
        next_num = 0
        num_count_list = list()
        for i in range(len(df)): # Check is CM number is already assigned and get the highest value
            identifier = df.loc[i, "Generator Function"]
            if identifier == "MTR" and pd.notna(df.loc[i, "CM No"]):
                num_count = int(df.loc[i, "CM No"])
                num_count_list.append(num_count)

        if num_count_list == []: # Assign starting CM number
            next_num = 2000
        else:
            next_num = max(num_count_list)

        for i in range(len(df)):
            identifier = df.loc[i, "Generator Function"]
            tag_part_3 = df.loc[i, "TAG Part 3 (Function code)"]
            if identifier == "MTR" and pd.isna(df.loc[i, "CM No"]):
                next_num = next_num + 1
                df.loc[i, "CM No"] = str(next_num)
                for x in range(len(df)):
                    identifier = df.loc[x, "Generator Function"]
                    if (identifier == "MTR_Dis") and (df.loc[x, "TAG Part 3 (Function code)"] == tag_part_3):
                        df.loc[x, "CM No"] = str(next_num)
                    elif (identifier == "MTR_Pro") and (df.loc[x, "TAG Part 3 (Function code)"] == tag_part_3):
                        df.loc[x, "CM No"] = str(next_num)
                    elif (identifier == "MTR_Out") and (df.loc[x, "TAG Part 3 (Function code)"] == tag_part_3):
                        df.loc[x, "CM No"] = str(next_num)
                    elif (identifier == "MTR_Run") and (df.loc[x, "TAG Part 3 (Function code)"] == tag_part_3):
                        df.loc[x, "CM No"] = str(next_num)
                    elif (identifier == "MTR_Rot") and (df.loc[x, "TAG Part 3 (Function code)"] == tag_part_3):
                        df.loc[x, "CM No"] = str(next_num)
                    elif (identifier == "MTR_Saf") and (df.loc[x, "TAG Part 3 (Function code)"] == tag_part_3):
                        df.loc[x, "CM No"] = str(next_num)
                    elif (identifier == "MTR_Saf_Fb") and (df.loc[x, "TAG Part 3 (Function code)"] == tag_part_3):
                        df.loc[x, "CM No"] = str(next_num)

    ##### AIN #####
        num_count = 0
        next_num = 0
        num_count_list = list()
        for i in range(len(df)): # Check is CM number is already assigned and get the highest value
            identifier = df.loc[i, "Generator Function"]
            if identifier == "AIN" and pd.notna(df.loc[i, "CM No"]):
                num_count = int(df.loc[i, "CM No"])
                num_count_list.append(num_count)

        if num_count_list == []: # Assign starting CM number
            next_num = 3000
        else:
            next_num = max(num_count_list)

        for i in range(len(df)):
            identifier = df.loc[i, "Generator Function"]
            if identifier == "AIN" and pd.isna(df.loc[i, "CM No"]):
                next_num = next_num + 1
                df.loc[i, "CM No"] = str(next_num)

    ##### AOUT #####
        num_count = 0
        next_num = 0
        num_count_list = list()
        for i in range(len(df)): # Check is CM number is already assigned and get the highest value
            identifier = df.loc[i, "Generator Function"]
            if identifier == "AOUT" and pd.notna(df.loc[i, "CM No"]):
                num_count = int(df.loc[i, "CM No"])
                num_count_list.append(num_count)

        if num_count_list == []: # Assign starting CM number
            next_num = 4000
        else:
            next_num = max(num_count_list)

        for i in range(len(df)):
            identifier = df.loc[i, "Generator Function"]
            if identifier == "AOUT" and pd.isna(df.loc[i, "CM No"]):
                next_num = next_num + 1
                df.loc[i, "CM No"] = str(next_num)

    ##### DIN #####
        num_count = 0
        next_num = 0
        num_count_list = list()
        for i in range(len(df)): # Check is CM number is already assigned and get the highest value
            identifier = df.loc[i, "Generator Function"]
            if identifier == "DIN" and pd.notna(df.loc[i, "CM No"]):
                num_count = int(df.loc[i, "CM No"])
                num_count_list.append(num_count)

        if num_count_list == []: # Assign starting CM number
            next_num = 5000
        else:
            next_num = max(num_count_list)

        for i in range(len(df)):
            identifier = df.loc[i, "Generator Function"]
            if identifier == "DIN" and pd.isna(df.loc[i, "CM No"]):
                next_num = next_num + 1
                df.loc[i, "CM No"] = str(next_num)

    ##### PID #####
        num_count = 0
        next_num = 0
        num_count_list = list()
        for i in range(len(df)): # Check is CM number is already assigned and get the highest value
            identifier = df.loc[i, "Generator Function"]
            if identifier == "PID" and pd.notna(df.loc[i, "CM No"]):
                num_count = int(df.loc[i, "CM No"])
                num_count_list.append(num_count)

        if num_count_list == []: # Assign starting CM number
            next_num = 6000
        else:
            next_num = max(num_count_list)

        for i in range(len(df)):
            identifier = df.loc[i, "Generator Function"]
            if identifier == "PID" and pd.isna(df.loc[i, "CM No"]):
                next_num = next_num + 1
                df.loc[i, "CM No"] = str(next_num)

    ##### ALM #####
        num_count = 0
        next_num = 0
        num_count_list = list()
        for i in range(len(df)): # Check is CM number is already assigned and get the highest value
            identifier = df.loc[i, "Generator Function"]
            if identifier == "ALM" and pd.notna(df.loc[i, "CM No"]):
                num_count = int(df.loc[i, "CM No"])
                num_count_list.append(num_count)

        if num_count_list == []: # Assign starting CM number
            next_num = 10000
        else:
            next_num = max(num_count_list)
    
        for i in range(len(df)):
            identifier = df.loc[i, "Generator Function"]
            if identifier == "ALM" and pd.isna(df.loc[i, "CM No"]):
                next_num = next_num + 1
                df.loc[i, "CM No"] = str(next_num)

    ##### VLV #####
        num_count = 0
        next_num = 0
        num_count_list = list()
        for i in range(len(df)): # Check is CM number is already assigned and get the highest value
            identifier = df.loc[i, "Generator Function"]
            if identifier == "VLV" and pd.notna(df.loc[i, "CM No"]):
                num_count = int(df.loc[i, "CM No"])
                num_count_list.append(num_count)

        if num_count_list == []: # Assign starting CM number
            next_num = 7000
        else:
            next_num = max(num_count_list)

        for i in range(len(df)):
            identifier = df.loc[i, "Generator Function"]
            tag_part_3 = df.loc[i, "TAG Part 3 (Function code)"]
            if identifier == "VLV" and pd.isna(df.loc[i, "CM No"]):
                next_num = next_num + 1
                df.loc[i, "CM No"] = str(next_num)
                for x in range(len(df)):
                    identifier = df.loc[x,"Generator Function"]
                    if (identifier == "VLV_N") and (df.loc[x, "TAG Part 3 (Function code)"] == tag_part_3):
                        df.loc[x, "CM No"] = str(next_num)
                    elif (identifier == "VLV_Out_N") and (df.loc[x, "TAG Part 3 (Function code)"] == tag_part_3):
                        df.loc[x, "CM No"] = str(next_num)
                    elif (identifier == "VLV_Out_P") and (df.loc[x, "TAG Part 3 (Function code)"] == tag_part_3):
                        df.loc[x, "CM No"] = str(next_num)
                    elif (identifier == "VLV_P") and (df.loc[x, "TAG Part 3 (Function code)"] == tag_part_3):
                        df.loc[x, "CM No"] = str(next_num)

    ##### SUM #####
        num_count = 0
        next_num = 0
        num_count_list = list()
        for i in range(len(df)): # Check is CM number is already assigned and get the highest value
            identifier = df.loc[i, "Generator Function"]
            if identifier == "SUM" and pd.notna(df.loc[i, "CM No"]):
                num_count = int(df.loc[i, "CM No"])
                num_count_list.append(num_count)

        if num_count_list == []: # Assign starting CM number
            next_num = 8000
        else:
            next_num = max(num_count_list)

        for i in range(len(df)):
            identifier = df.loc[i, "Generator Function"]
            if identifier == "SUM" and pd.isna(df.loc[i, "CM No"]):
                next_num = next_num + 1
                df.loc[i, "CM No"] = str(next_num)

    def CreateFRQSheet(frq_list): # ----------------Sheet for FRQ component----------------
        
        for i in range(len(df)):
            if str(df.loc[i, "Sort"]) == "1" and pd.notna(df.loc[i, "Component Description"]): area = df.loc[i, "Component Description"]
            identifier = df.loc[i,"Generator Function"]
            if identifier == ("FRQ"):
                    tagPart2 = df.loc[i,"TAG Part 2 (# Item)"]
                    tagPart2_noDot = tagPart2.split('.', 1)[0]
                    frq_dict = {'Tag Part 2'                : df.loc[i, "TAG Part 2 (# Item)"],
                                'Tag Part 3'                : df.loc[i, "TAG Part 3 (Function code)"],
                                'Tag Part 4'                : df.loc[i, "TAG Part 4 (# Function serial)"],
                                'Tag'                       : f'{tagPart2_noDot}_{df.loc[i, "TAG Part 3 (Function code)"]}_{df.loc[i, "TAG Part 4 (# Function serial)"]}',
                                'Description'               : f'{df.loc[i, "Component Description"]} - {df.loc[i, "Process Description"]}',
                                'CM No'                     : df.loc[i, "CM No"],
                                'Alarm Group'               : '',
                                'Cabinet'                   : df.loc[i, "TAG Part 2 (# Item)"],
                                'Area'                      : area,
                                'FRQ Type'                  : '',
                                'Thermal Fault'             : '',
                                'Rotation Sensor'           : '',
                                'Thermistor'                : '',
                                'Tab ID'                    : df.loc[i, "Tab ID"],
                                'Sub ID'                    : df.loc[i, "Sub ID"],
                                'SC Local'                  : df.loc[i, "SC Local"],
                                'Main Power [KW]'           : df.loc[i, "Main Power [kW]"],
                                'Main Current [A]'          : df.loc[i, "Main Current [A]"],
                                'Main Voltage [V]'          : df.loc[i, "Main Voltage [V]"],
                                'Connection'                : df.loc[i, "Control Current [A] OR [Nm] or (Connection)"],
                                'Poles'                     : df.loc[i, "Poles"],
                                'Frequency [Hz]'            : df.loc[i, "Frequency [Hz]"],
                                'Cos φ'                     : df.loc[i, "Cos φ (Efficiency)"],
                                'RPM'                       : df.loc[i, "RPM"],
                                'EU'                        : df.loc[i, "EU"],
                                'Min EU'                    : df.loc[i, "Min EU"],
                                'Max EU'                    : df.loc[i, "Max EU"],
                                'Raw Min'                   : df.loc[i, "Raw Min"],
                                'Raw Max'                   : df.loc[i, "Raw Max"],
                                'PLC No.'                   : df.loc[i, "PLC No"],
                                'Disconnect'                : '',
                                'IP Address'                : '',
                                'Electrical Room'           : '',
                                'PLC -> SCADA'              : f'{{[plc{df.loc[i, "PLC No"]}]{tagPart2_noDot}_{df.loc[i, "TAG Part 3 (Function code)"]}_{df.loc[i, "TAG Part 4 (# Function serial)"]}_{identifier}_UDT}}',
                                'SCADA Tag'                 : f'{{{tagPart2_noDot}_{df.loc[i, "TAG Part 3 (Function code)"]}_{df.loc[i, "TAG Part 4 (# Function serial)"]}_{identifier}}}',
                                'Checked'                   : ''
                                }
                    frq_list.append(frq_dict)
                

    # Search the list for a dict with a matching generator config, and modify the corresponding dict.
        for i in range(len(df)):
            identifier = df.loc[i, "Generator Function"]
            if identifier == ("FRQ_Pro"):
                gen_con = df.loc[i, "CM No"]
                list_index = next((index for (index, d) in enumerate(frq_list) if d["CM No"] == gen_con), None)
                if list_index is not None:
                    frq_list[int(list_index)]['Thermal Fault'] = df.loc[i, "PLC Tag Name"]
        for i in range(len(df)):
            identifier = df.loc[i, "Generator Function"]
            if identifier == ("FRQ_Rot"):
                gen_con = df.loc[i, "CM No"]
                list_index = next((index for (index, d) in enumerate(frq_list) if d["CM No"] == gen_con), None)
                if list_index is not None:
                    frq_list[int(list_index)]['Rotation Sensor'] = df.loc[i, "PLC Tag Name"]
        for i in range(len(df)):
            identifier = df.loc[i, "Generator Function"]
            if identifier == ("FRQ_Ptc"):
                gen_con = df.loc[i, "CM No"]
                list_index = next((index for (index, d) in enumerate(frq_list) if d["CM No"] == gen_con), None)
                if list_index is not None:
                    frq_list[int(list_index)]['Thermistor'] = df.loc[i, "Process Description"]
        for i in range(len(df)):
            identifier = df.loc[i, "Generator Function"]
            if identifier == ("FRQ_Dis"):
                gen_con = df.loc[i, "CM No"]
                list_index = next((index for (index, d) in enumerate(frq_list) if d["CM No"] == gen_con), None)
                if list_index is not None:
                    frq_list[int(list_index)]['Disconnect'] = df.loc[i, "PLC Tag Name"]
        for i in range(len(df)):
            identifier = df.loc[i, "Generator Function"]
            if identifier == ("FRQ_ETH"):
                gen_con = df.loc[i, "CM No"]
                list_index = next((index for (index, d) in enumerate(frq_list) if d["CM No"] == gen_con), None)
                if list_index is not None:
                    frq_list[int(list_index)]['IP Address'] = df.loc[i, "Net number / IP address"]
                    frq_list[int(list_index)]['FRQ Type'] = df.loc[i, "FRQ Type"]
                    frq_list[int(list_index)]['Electrical Room'] = df.loc[i, "TAG Part 1 (# System)"]
        return frq_list
    

    def CreateMTRSheet(mtr_list): # ----------------Sheet for MTR component----------------
    
        for i in range(len(df)):
            if str(df.loc[i, "Sort"]) == "1" and pd.notna(df.loc[i, "Component Description"]): area = df.loc[i, "Component Description"]
            identifier = df.loc[i, "Generator Function"]
            if identifier == ("MTR"):
                    tagPart2 = df.loc[i, "TAG Part 2 (# Item)"]
                    tagPart2_noDot = tagPart2.split('.', 1)[0]
                    mtr_dict = {'Tag Part 2'                    : df.loc[i, "TAG Part 2 (# Item)"],
                                'Tag Part 3'                    : df.loc[i, "TAG Part 3 (Function code)"],
                                'Tag Part 4'                    : df.loc[i, "TAG Part 4 (# Function serial)"],
                                'Tag'                           : f'{tagPart2_noDot}_{df.loc[i, "TAG Part 3 (Function code)"]}_{df.loc[i, "TAG Part 4 (# Function serial)"]}',
                                'Description'                   : f'{df.loc[i, "Component Description"]} - {df.loc[i, "Process Description"]}',
                                'CM No'                         : df.loc[i, "CM No"],
                                'Alarm Group'                   : '',
                                'Cabinet'                       : df.loc[i, "TAG Part 2 (# Item)"],
                                'Area'                          : area,
                                'Tab ID'                        : df.loc[i, "Tab ID"],
                                'Sub ID'                        : df.loc[i, "Sub ID"],
                                'SC Local'                      : df.loc[i, "SC Local"],
                                'Thermal Fault'                 : '',
                                'Rotation Sensor'               : '',
                                'Thermistor'                    : '',
                                'Start Forward'                 : '',
                                'Start Reverse'                 : '',
                                'Forward Feedback'              : '',
                                'Reverse Feedback'              : '',
                                'Emergency Stop Contactor Fb'   : '',
                                'Emergency stop Relay Fb'       : '',
                                'PLC No.'                       : df.loc[i, "PLC No"],
                                'Emergency Stop Contactor'      : '',
                                'Emergency Stop Contactor FB'   : '',
                                'Disconnect'                    : '',
                                'Forward PLC address'           : f'IP:{df.loc[i, "Net number / IP address"]}', #Only when IP connection to MTR is used else ('',)
                                'Electrical Room'               : '',
                                'PLC -> SCADA'                  : f'{{[plc{df.loc[i, "PLC No"]}]{tagPart2_noDot}_{df.loc[i, "TAG Part 3 (Function code)"]}_{df.loc[i, "TAG Part 4 (# Function serial)"]}_{identifier}_UDT}}',
                                'SCADA Tag'                     : f'{{{tagPart2_noDot}_{df.loc[i, "TAG Part 3 (Function code)"]}_{df.loc[i, "TAG Part 4 (# Function serial)"]}_{identifier}}}',
                                'Checked'                       : ''
                                }
                    mtr_list.append(mtr_dict)

    # Search the list for a dict with a matching generator config, and modify the corresponding dict.
        for i in range(len(df)):
            identifier = df.loc[i, "Generator Function"]
            if identifier == ("MTR_Dis"):
                gen_con = df.loc[i, "CM No"]
                list_index = next((index for (index, d) in enumerate(mtr_list) if d["CM No"] == gen_con), None)
                if list_index is not None:
                    mtr_list[int(list_index)]['Disconnect'] = df.loc[i, "PLC Tag Name"]
        for i in range(len(df)):
            identifier = df.loc[i, "Generator Function"]
            if identifier == ("MTR_Pro"):
                gen_con = df.loc[i, "CM No"]
                list_index = next((index for (index, d) in enumerate(mtr_list) if d["CM No"] == gen_con), None)
                if list_index is not None:
                    mtr_list[int(list_index)]['Thermal Fault'] = df.loc[i, "PLC Tag Name"]
                    mtr_list[int(list_index)]['Electrical Room'] = df.loc[i, "TAG Part 1 (# System)"]
        for i in range(len(df)):
            identifier = df.loc[i, "Generator Function"]
            if identifier == ("MTR_Rot"):
                gen_con = df.loc[i, "CM No"]
                list_index = next((index for (index, d) in enumerate(mtr_list) if d["CM No"] == gen_con), None)
                if list_index is not None:
                    mtr_list[int(list_index)]['Rotation Sensor'] = df.loc[i, "PLC Tag Name"]
        for i in range(len(df)):
            identifier = df.loc[i, "Generator Function"]
            if identifier == ("MTR_Out"):
                gen_con = df.loc[i, "CM No"]
                list_index = next((index for (index, d) in enumerate(mtr_list) if d["CM No"] == gen_con), None)
                if list_index is not None:
                    mtr_list[int(list_index)]['Start Forward'] = df.loc[i, "PLC Tag Name"]
                    mtr_list[int(list_index)]['Forward PLC address'] = df.loc[i, "PLC Address"]
        for i in range(len(df)):
            identifier = df.loc[i, "Generator Function"]
            if identifier == ("MTR_Out_Rev"):
                gen_con = df.loc[i, "CM No"]
                list_index = next((index for (index, d) in enumerate(mtr_list) if d["CM No"] == gen_con), None)
                if list_index is not None:
                    mtr_list[int(list_index)]['Start Reverse'] = df.loc[i, "PLC Tag Name"]
        for i in range(len(df)):
            identifier = df.loc[i, "Generator Function"]
            if identifier == ("MTR_Saf"):
                gen_con = df.loc[i, "CM No"]
                list_index = next((index for (index, d) in enumerate(mtr_list) if d["CM No"] == gen_con), None)
                if list_index is not None:
                    mtr_list[int(list_index)]['Emergency Stop Contactor Fb'] = df.loc[i, "PLC Tag Name"]
        for i in range(len(df)):
            identifier = df.loc[i, "Generator Function"]
            if identifier == ("ALM"):
                gen_con = df.loc[i, "CM No"]
                list_index = next((index for (index, d) in enumerate(mtr_list) if d["CM No"] == gen_con), None)
                if list_index is not None:
                    mtr_list[int(list_index)]['Emergency stop Relay Fb'] = df.loc[i, "PLC Tag Name"]
        for i in range(len(df)):
            identifier = df.loc[i, "Generator Function"]
            if identifier == ("MTR_Run"):
                gen_con = df.loc[i, "CM No"]
                list_index = next((index for (index, d) in enumerate(mtr_list) if d["CM No"] == gen_con), None)
                if list_index is not None:
                    mtr_list[int(list_index)]['Forward Feedback'] = df.loc[i, "PLC Tag Name"]
        for i in range(len(df)):
            identifier = df.loc[i, "Generator Function"]
            if identifier == ("MTR_Run_Rev"):
                gen_con = df.loc[i, "CM No"]
                list_index = next((index for (index, d) in enumerate(mtr_list) if d["CM No"] == gen_con), None)
                if list_index is not None:
                    mtr_list[int(list_index)]['Reverse Feedback'] = df.loc[i, "PLC Tag Name"]
        for i in range(len(df)):
            identifier = df.loc[i, "Generator Function"]
            if identifier == ("MTR_Saf"):
                gen_con = df.loc[i, "CM No"]
                list_index = next((index for (index, d) in enumerate(mtr_list) if d["CM No"] == gen_con), None)
                if list_index is not None:
                    mtr_list[int(list_index)]['Emergency Stop Contactor'] = df.loc[i, "PLC Tag Name"]
        for i in range(len(df)):
            identifier = df.loc[i, "Generator Function"]
            if identifier == ("MTR_Saf_Fb"):
                gen_con = df.loc[i, "CM No"]
                list_index = next((index for (index, d) in enumerate(mtr_list) if d["CM No"] == gen_con), None)
                if list_index is not None:
                    mtr_list[int(list_index)]['Emergency Stop Contactor FB'] = df.loc[i, "PLC Tag Name"]

        return mtr_list

    def CreateVLVSheet(vlv_list): # ----------------Sheet for VLV component---------------

        for i in range(len(df)):
            if str(df.loc[i, "Sort"]) == "1" and pd.notna(df.loc[i, "Component Description"]): area = df.loc[i, "Component Description"]
            identifier = str(df.loc[i, "Generator Function"])
            if identifier == ("VLV"):
                    tagPart2 = df.loc[i, "TAG Part 2 (# Item)"]
                    tagPart2_noDot = tagPart2.split('.', 1)[0]
                    vlv_dict = {'Tag Part 2'                : df.loc[i, "TAG Part 2 (# Item)"],
                                'Tag Part 3'                : df.loc[i, "TAG Part 3 (Function code)"],
                                'Tag Part 4'                : df.loc[i, "TAG Part 4 (# Function serial)"],
                                'Tag'                       : f'{tagPart2_noDot}_{df.loc[i, "TAG Part 3 (Function code)"]}_{df.loc[i, "TAG Part 4 (# Function serial)"]}',
                                'Description'               : f'{df.loc[i, "Component Description"]} - {df.loc[i, "Process Description"]}',
                                'CM No'                     : df.loc[i, "CM No"],
                                'IO-address'                : '',
                                'Cabinet'                   : df.loc[i, "TAG Part 2 (# Item)"],
                                'Area'                      : area,
                                'Tab ID'                    : df.loc[i, "Tab ID"],
                                'Sub ID'                    : df.loc[i, "Sub ID"],
                                'SC Local'                  : df.loc[i, "SC Local"],
                                'Valve Out P'               : '',
                                'Valve Out N'               : '',
                                'Valve FB P'                : '',
                                'Valve FB N'                : '',
                                'Activate Text'             : df.loc[i, "VLV button P-Activate text"],
                                'Deactivate Text'           : df.loc[i, "VLV button N-DeAct text"],
                                'PLC No.'                   : df.loc[i, "PLC No"],
                                'PLC -> SCADA'              : f'{{[plc{df.loc[i, "PLC No"]}]{tagPart2_noDot}_{df.loc[i, "TAG Part 3 (Function code)"]}_{df.loc[i, "TAG Part 4 (# Function serial)"]}_{identifier}_UDT}}',
                                'SCADA Tag'                 : f'{{{tagPart2_noDot}_{df.loc[i, "TAG Part 3 (Function code)"]}_{df.loc[i, "TAG Part 4 (# Function serial)"]}_{identifier}}}',
                                'Checked'                   : ''
                                }
                    vlv_list.append(vlv_dict)

        for i in range(len(df)):
            identifier = df.loc[i, "Generator Function"]
            if identifier == ("VLV_Out_P"):
                gen_con = df.loc[i, "CM No"]
                list_index = next((index for (index, d) in enumerate(vlv_list) if d["CM No"] == gen_con), None)
                if list_index is not None:
                    vlv_list[int(list_index)]['Valve Out P'] = df.loc[i, "PLC Tag Name"]
                    vlv_list[int(list_index)]['IO-address'] = df.loc[i, "PLC Address"]
        for i in range(len(df)):
            identifier = df.loc[i, "Generator Function"]
            if identifier == ("VLV_Out_N"):
                gen_con = df.loc[i, "CM No"]
                list_index = next((index for (index, d) in enumerate(vlv_list) if d["CM No"] == gen_con), None)
                if list_index is not None:
                    vlv_list[int(list_index)]['Valve Out N'] = df.loc[i, "PLC Tag Name"]
        for i in range(len(df)):
            identifier = df.loc[i, "Generator Function"]
            if identifier == ("VLV_P"):
                gen_con = df.loc[i, "CM No"]
                list_index = next((index for (index, d) in enumerate(vlv_list) if d["CM No"] == gen_con), None)
                if list_index is not None:
                    vlv_list[int(list_index)]['Valve FB P'] = df.loc[i, "PLC Tag Name"]
        for i in range(len(df)):
            identifier = df.loc[i, "Generator Function"]
            if identifier == ("VLV_N"):
                gen_con = df.loc[i, "CM No"]
                list_index = next((index for (index, d) in enumerate(vlv_list) if d["CM No"] == gen_con), None)
                if list_index is not None:
                    vlv_list[int(list_index)]['Valve FB N'] = df.loc[i, "PLC Tag Name"]

        return vlv_list

    def CreateDINSheet(din_list): # ----------------Sheet for DIN component----------------

        for i in range(len(df)):
            if str(df.loc[i, "Sort"]) == "1" and pd.notna(df.loc[i, "Component Description"]): area = df.loc[i, "Component Description"]
            identifier = str(df.loc[i, "Generator Function"])
            if identifier == ("DIN"):
                    tagPart2 = df.loc[i, "TAG Part 2 (# Item)"]
                    tagPart2_noDot = str(tagPart2.split('.', 1)[0])
                    din_dict = {'Tag Part 2'                : str(df.loc[i, "TAG Part 2 (# Item)"]),
                                'Tag Part 3'                : str(df.loc[i, "TAG Part 3 (Function code)"]),
                                'Tag Part 4'                : str(df.loc[i, "TAG Part 4 (# Function serial)"]),
                                'Tag'                       : f'{tagPart2_noDot}_{df.loc[i, "TAG Part 3 (Function code)"]}_{df.loc[i, "TAG Part 4 (# Function serial)"]}',
                                'Description'               : f'{df.loc[i, "Component Description"]} - {df.loc[i, "Process Description"]}',
                                'CM No'                     : str(df.loc[i, "CM No"]),
                                'Alarm Group'               : '',
                                'Cabinet'                   : str(df.loc[i, "TAG Part 2 (# Item)"]),
                                'Area'                      : str(area),
                                'Tab ID'                    : str(df.loc[i, "Tab ID"]),
                                'Sub ID'                    : str(df.loc[i, "Sub ID"]),
                                'SC Local'                  : str(df.loc[i, "SC Local"]),
                                'PLC Tag Name'              : str(df.loc[i, "PLC Tag Name"]),
                                'PLC No.'                   : str(df.loc[i, "PLC No"]),
                                'T. Funk'                   : str(df.loc[i, "PLC Address"]),
                                'PLC -> SCADA'              : f'{{[plc{df.loc[i, "PLC No"]}]{tagPart2_noDot}_{df.loc[i, "TAG Part 3 (Function code)"]}_{df.loc[i, "TAG Part 4 (# Function serial)"]}_{identifier}_UDT}}',
                                'SCADA Tag'                 : f'{{{tagPart2_noDot}_{df.loc[i, "TAG Part 3 (Function code)"]}_{df.loc[i, "TAG Part 4 (# Function serial)"]}_{identifier}}}',
                                'Checked'                   : ''
                                }
                    din_list.append(din_dict)
        return din_list

    def CreateAINSheet(ain_list): # ----------------Sheet for AIN component----------------

        for i in range(len(df)):
            if str(df.loc[i, "Sort"]) == "1" and pd.notna(df.loc[i, "Component Description"]): area = df.loc[i, "Component Description"]
            identifier = df.loc[i, "Generator Function"]
            if identifier == ("AIN"):
                    tagPart2 = df.loc[i, "TAG Part 2 (# Item)"]
                    tagPart2_noDot = tagPart2.split('.', 1)[0]
                    ain_dict = {'Tag Part 2'                : df.loc[i, "TAG Part 2 (# Item)"],
                                'Tag Part 3'                : df.loc[i, "TAG Part 3 (Function code)"],
                                'Tag Part 4'                : df.loc[i, "TAG Part 4 (# Function serial)"],
                                'Tag'                       : f'{tagPart2_noDot}_{df.loc[i, "TAG Part 3 (Function code)"]}_{df.loc[i, "TAG Part 4 (# Function serial)"]}',
                                'Description'               : f'{df.loc[i, "Component Description"]} - {df.loc[i, "Process Description"]}',
                                'CM No'                     : df.loc[i, "CM No"],
                                'Alarm Group'               : '',
                                'Cabinet'                   : df.loc[i, "TAG Part 2 (# Item)"],
                                'Area'                      : area,
                                'Tab ID'                    : df.loc[i, "Tab ID"],
                                'Sub ID'                    : df.loc[i, "Sub ID"],
                                'SC Local'                  : df.loc[i, "SC Local"],
                                'PLC Tag Name'              : df.loc[i, "PLC Tag Name"],
                                'PLC No.'                   : df.loc[i, "PLC No"],
                                'T. Funk'                   : df.loc[i, "PLC Address"],
                                'EU'                        : df.loc[i, "EU"],
                                'PLC -> SCADA'              : f'{{[plc{df.loc[i, "PLC No"]}]{tagPart2_noDot}_{df.loc[i, "TAG Part 3 (Function code)"]}_{df.loc[i, "TAG Part 4 (# Function serial)"]}_{identifier}_UDT}}',
                                'SCADA Tag'                 : f'{{{tagPart2_noDot}_{df.loc[i, "TAG Part 3 (Function code)"]}_{df.loc[i, "TAG Part 4 (# Function serial)"]}_{identifier}}}',
                                'Checked'                   : ''
                                }
                    ain_list.append(ain_dict)
        return ain_list

    def CreateAOUTSheet(aout_list): # ----------------Sheet for AOUT component----------------

        for i in range(len(df)):
            if str(df.loc[i, "Sort"]) == "1" and pd.notna(df.loc[i, "Component Description"]): area = df.loc[i, "Component Description"]
            identifier = df.loc[i, "Generator Function"]
            if identifier == ("AOUT"):
                    tagPart2 = df.loc[i, "TAG Part 2 (# Item)"]
                    tagPart2_noDot = tagPart2.split('.', 1)[0]
                    aout_dict = {'Tag Part 2'               : df.loc[i, "TAG Part 2 (# Item)"],
                                'Tag Part 3'                : df.loc[i, "TAG Part 3 (Function code)"],
                                'Tag Part 4'                : df.loc[i, "TAG Part 4 (# Function serial)"],
                                'Tag'                       : f'{tagPart2_noDot}_{df.loc[i, "TAG Part 3 (Function code)"]}_{df.loc[i, "TAG Part 4 (# Function serial)"]}',
                                'Description'               : f'{df.loc[i, "Component Description"]} - {df.loc[i, "Process Description"]}',
                                'CM No'                     : df.loc[i, "CM No"],
                                'Alarm Group'               : '',
                                'Cabinet'                   : df.loc[i, "TAG Part 2 (# Item)"],
                                'Area'                      : area,
                                'Tab ID'                    : df.loc[i, "Tab ID"],
                                'Sub ID'                    : df.loc[i, "Sub ID"],
                                'SC Local'                  : df.loc[i, "SC Local"],
                                'PLC Tag Name'              : df.loc[i, "PLC Tag Name"],
                                'PLC No.'                   : df.loc[i, "PLC No"],
                                'T. Funk'                   : df.loc[i, "PLC Address"],
                                'EU'                        : df.loc[i, "EU"],
                                'PLC -> SCADA'              : f'{{[plc{df.loc[i, "PLC No"]}]{tagPart2_noDot}_{df.loc[i, "TAG Part 3 (Function code)"]}_{df.loc[i, "TAG Part 4 (# Function serial)"]}_{identifier}_UDT}}',
                                'SCADA Tag'                 : f'{{{tagPart2_noDot}_{df.loc[i, "TAG Part 3 (Function code)"]}_{df.loc[i, "TAG Part 4 (# Function serial)"]}_{identifier}}}',
                                'Checked'                   : ''
                                }
                    aout_list.append(aout_dict)
        return aout_list
    
    def CreatePIDSheet(pid_list): # ----------------Sheet for PID component----------------

        for i in range(len(df)):
            if str(df.loc[i, "Sort"]) == "1" and pd.notna(df.loc[i, "Component Description"]): area = df.loc[i, "Component Description"]
            identifier = df.loc[i, "Generator Function"]
            if identifier == ("PID"):
                    tagPart2 = df.loc[i, "TAG Part 2 (# Item)"]
                    tagPart2_noDot = tagPart2.split('.', 1)[0]
                    pid_dict = {'Tag Part 2'                : df.loc[i, "TAG Part 2 (# Item)"],
                                'Tag Part 3'                : df.loc[i, "TAG Part 3 (Function code)"],
                                'Tag Part 4'                : df.loc[i, "TAG Part 4 (# Function serial)"],
                                'Tag'                       : f'{tagPart2_noDot}_{df.loc[i, "TAG Part 3 (Function code)"]}_{df.loc[i, "TAG Part 4 (# Function serial)"]}',
                                'Description'               : f'{df.loc[i, "Component Description"]} - {df.loc[i, "Process Description"]}',
                                'CM No'                     : df.loc[i, "CM No"],
                                'Alarm Group'               : '',
                                'Cabinet'                   : df.loc[i, "TAG Part 2 (# Item)"],
                                'Area'                      : area,
                                'Tab ID'                    : df.loc[i, "Tab ID"],
                                'Sub ID'                    : df.loc[i, "Sub ID"],
                                'SC Local'                  : df.loc[i, "SC Local"],
                                'PLC Tag Name'              : df.loc[i, "PLC Tag Name"],
                                'PLC No.'                   : str(df.loc[i, "PLC No"]),
                                'EU'                        : df.loc[i, "EU"],
                                'PLC -> SCADA'              : f'{{[plc{df.loc[i, "PLC No"]}]{tagPart2_noDot}_{df.loc[i, "TAG Part 3 (Function code)"]}_{df.loc[i, "TAG Part 4 (# Function serial)"]}_{identifier}_UDT}}',
                                'SCADA Tag'                 : f'{{{tagPart2_noDot}_{df.loc[i, "TAG Part 3 (Function code)"]}_{df.loc[i, "TAG Part 4 (# Function serial)"]}_{identifier}}}',
                                'Checked'                   : ''
                                }
                    pid_list.append(pid_dict)
        return pid_list

    def CreateALMSheet(alm_list): # ----------------Sheet for ALM component----------------

        for i in range(len(df)):
            if str(df.loc[i, "Sort"]) == "1" and pd.notna(df.loc[i, "Component Description"]): area = df.loc[i, "Component Description"]
            identifier = df.loc[i, "Generator Function"]
            if identifier == ("ALM"):
                    tagPart2 = df.loc[i, "TAG Part 2 (# Item)"]
                    tagPart2_noDot = tagPart2.split('.', 1)[0]
                    tagPart3 = df.loc[i, "TAG Part 3 (Function code)"]
                    tagPart3_noDot = tagPart3.split('.', 1)[0]
                    if str(df.loc[i, "Signal Type"]) == "ETH": plcTagName = f'{tagPart2_noDot}_{df.loc[i, "TAG Part 3 (Function code)"]}_{df.loc[i, "TAG Part 4 (# Function serial)"]}_ALM_UDT.S.Comm_Status'
                    else: plcTagName = df.loc[i, "PLC Tag Name"]
                    alm_dict = {'Tag Part 2'                : df.loc[i, "TAG Part 2 (# Item)"],
                                'Tag Part 3'                : df.loc[i, "TAG Part 3 (Function code)"],
                                'Tag Part 4'                : df.loc[i, "TAG Part 4 (# Function serial)"],
                                'Tag'                       : f'{tagPart2_noDot}_{tagPart3_noDot}_{df.loc[i, "TAG Part 4 (# Function serial)"]}',
                                'Description'               : f'{df.loc[i, "Component Description"]} - {df.loc[i, "Process Description"]}',
                                'CM No'                     : df.loc[i, "CM No"],
                                'Alarm Group'               : '',
                                'Cabinet'                   : df.loc[i, "TAG Part 2 (# Item)"],
                                'Area'                      : area,
                                'Tab ID'                    : df.loc[i, "Tab ID"],
                                'Sub ID'                    : df.loc[i, "Sub ID"],
                                'SC Local'                  : df.loc[i, "SC Local"],
                                'PLC Tag Name'              : f'{plcTagName}',
                                'PLC No.'                   : df.loc[i, "PLC No"],
                                'T. Funk'                   : df.loc[i, "PLC Address"],
                                'PLC -> SCADA'              : f'{{[plc{df.loc[i, "PLC No"]}]{tagPart2_noDot}_{df.loc[i, "TAG Part 3 (Function code)"]}_{df.loc[i, "TAG Part 4 (# Function serial)"]}_{identifier}_UDT}}',                                            
                                'SCADA Tag'                 : f'{{{tagPart2_noDot}_{df.loc[i, "TAG Part 3 (Function code)"]}_{df.loc[i, "TAG Part 4 (# Function serial)"]}_{identifier}}}',
                                'Checked'                   : ''
                                }
                    alm_list.append(alm_dict)
        return alm_list

    def CreateFPSheet(fp_list): # ----------------Sheet for FP component----------------

        for i in range(len(df)):
            if str(df.loc[i, "Sort"]) == "1" and pd.notna(df.loc[i, "Component Description"]): area = df.loc[i, "Component Description"]
            identifier = str(df.loc[i, "Generator Function"])
            if identifier in identifiers:
                    tagPart2 = df.loc[i, "TAG Part 2 (# Item)"]
                    tagPart2_noDot = tagPart2.split('.', 1)[0]
                    tagPart3 = df.loc[i, "TAG Part 3 (Function code)"]
                    tagPart3_noDot = tagPart3.split('.', 1)[0]
                    fp_dict = {'Tag Part 2'                 : df.loc[i, "TAG Part 2 (# Item)"],
                                'Tag Part 3'                : df.loc[i, "TAG Part 3 (Function code)"],
                                'Tag Part 4'                : df.loc[i, "TAG Part 4 (# Function serial)"],
                                'Tag'                       : f'{tagPart2_noDot}_{tagPart3_noDot}_{df.loc[i, "TAG Part 4 (# Function serial)"]}',
                                'Description'               : f'{df.loc[i, "Component Description"]} - {df.loc[i, "Process Description"]}',
                                'CM No'                     : df.loc[i, "CM No"],
                                'Alarm Group'               : '',
                                'Cabinet'                   : df.loc[i, "TAG Part 2 (# Item)"],
                                'Area'                      : area,
                                'Tab ID'                    : df.loc[i, "Tab ID"],
                                'Sub ID'                    : df.loc[i, "Sub ID"],
                                'SC Local'                  : df.loc[i, "SC Local"],
                                'EU'                        : df.loc[i, "EU"],
                                'PLC Address'               : df.loc[i, "PLC Address"],
                                'PLC No.'                   : df.loc[i, "PLC No"]
                                }
                    fp_list.append(fp_dict)
        return fp_list

    def CreateFPVLVSheet(fp_vlv_list): # ----------------Sheet for FP_VLV component----------------

        for i in range(len(df)):
            if str(df.loc[i, "Sort"]) == "1" and pd.notna(df.loc[i, "Component Description"]): area = df.loc[i, "Component Description"]
            identifier = str(df.loc[i, "Generator Function"])
            if identifier == ("VLV"):
                    tagPart2 = df.loc[i, "TAG Part 2 (# Item)"]
                    tagPart2_noDot = tagPart2.split('.', 1)[0]
                    fp_vlv_dict = {'Tag Part 2'                 : df.loc[i, "TAG Part 2 (# Item)"],
                                    'Tag Part 3'                : df.loc[i, "TAG Part 3 (Function code)"],
                                    'Tag Part 4'                : df.loc[i, "TAG Part 4 (# Function serial)"],
                                    'Tag'                       : f'{tagPart2_noDot}_{df.loc[i, "TAG Part 3 (Function code)"]}_{df.loc[i, "TAG Part 4 (# Function serial)"]}',
                                    'Description'               : f'{df.loc[i, "Component Description"]} - {df.loc[i, "Process Description"]}',
                                    'CM No'                     : df.loc[i, "CM No"],
                                    'Alarm Group'               : '',
                                    'Cabinet'                   : df.loc[i, "TAG Part 2 (# Item)"],
                                    'Area'                      : area,
                                    'Tab ID'                    : df.loc[i, "Tab ID"],
                                    'Sub ID'                    : df.loc[i, "Sub ID"],
                                    'SC Local'                  : df.loc[i, "SC Local"],
                                    'EU'                        : df.loc[i, "EU"],
                                    'PLC Address'               : df.loc[i, "PLC Address"],
                                    'Valve Out P'               : '',
                                    'Valve Out N'               : '',
                                    'Valve FB P'                : '',
                                    'Valve FB N'                : '',
                                    'Activate Text'             : df.loc[i, "VLV button P-Activate text"],
                                    'Deactivate Text'           : df.loc[i, "VLV button N-DeAct text"],
                                    'PLC No.'                   : df.loc[i, "PLC No"]
                                }
                    fp_vlv_list.append(fp_vlv_dict)
        return fp_vlv_list


    def CreateSAFESheet(aoi_safe_list): # ----------------Sheet for SAFE component----------------

        for i in range(len(df)):
            if str(df.loc[i, "Sort"]) == "1" and pd.notna(df.loc[i, "Component Description"]): area = df.loc[i, "Component Description"]
            identifier = df.loc[i, "Generator Function"]
            signal_type1 = df.loc[i, "Signal Type"] == "SDI"
            signal_type2 = df.loc[i, "Signal Type"] == "SDO"
            sort_nr1 = df.loc[i, "Sort"] == 3
            sort_nr2 = df.loc[i, "Sort"] == "3"
            if identifier in identifiers and (signal_type1 or signal_type2) and (sort_nr1 or sort_nr2):
                    tagPart2 = df.loc[i, "TAG Part 2 (# Item)"]
                    tagPart2_noDot = tagPart2.split('.', 1)[0]
                    aoi_safe_dict = {'Tag Part 2'               : df.loc[i, "TAG Part 2 (# Item)"],
                                    'Tag Part 3'                : df.loc[i, "TAG Part 3 (Function code)"],
                                    'Tag Part 4'                : df.loc[i, "TAG Part 4 (# Function serial)"],
                                    'Tag'                       : f'{tagPart2_noDot}_{df.loc[i, "TAG Part 3 (Function code)"]}_{df.loc[i, "TAG Part 4 (# Function serial)"]}',
                                    'Description'               : f'{df.loc[i, "Component Description"]} - {df.loc[i, "Process Description"]}',
                                    'CM No'                     : df.loc[i, "CM No"],
                                    'Alarm Group'               : '',
                                    'Cabinet'                   : df.loc[i, "TAG Part 2 (# Item)"],
                                    'Area'                      : area,
                                    'Tab ID'                    : df.loc[i, "Tab ID"],
                                    'Sub ID'                    : df.loc[i, "Sub ID"],
                                    'SC Local'                  : df.loc[i, "SC Local"],
                                    'PLC No.'                   : df.loc[i, "PLC No"],
                                    'T. Funk'                   : df.loc[i, "PLC Address"]
                                }
                    aoi_safe_list.append(aoi_safe_dict)
        return aoi_safe_list


    def CreateSUMSheet(sum_list): # ----------------Sheet for SUM - Power Monitor component----------------

        for i in range(len(df)):
            if str(df.loc[i, "Sort"]) == "1" and pd.notna(df.loc[i, "Component Description"]): area = df.loc[i, "Component Description"]
            identifier = df.loc[i, "Generator Function"]
            if identifier == ("SUM"):
                    tagPart2 = df.loc[i, "TAG Part 2 (# Item)"]
                    tagPart2_noDot = tagPart2.split('.', 1)[0]
                    sum_dict = {'Tag Part 2'                    : df.loc[i, "TAG Part 2 (# Item)"],
                                    'Tag Part 3'                : df.loc[i, "TAG Part 3 (Function code)"],
                                    'Tag Part 4'                : df.loc[i, "TAG Part 4 (# Function serial)"],
                                    'Tag'                       : f'{tagPart2_noDot}_{df.loc[i, "TAG Part 3 (Function code)"]}_{df.loc[i, "TAG Part 4 (# Function serial)"]}',
                                    'Description'               : f'{df.loc[i, "Component Description"]} - {df.loc[i, "Process Description"]}',
                                    'CM No'                     : df.loc[i, "CM No"],
                                    'Alarm Group'               : '',
                                    'Cabinet'                   : df.loc[i, "TAG Part 2 (# Item)"],
                                    'Area'                      : area,
                                    'Tab ID'                    : df.loc[i, "Tab ID"],
                                    'Sub ID'                    : df.loc[i, "Sub ID"],
                                    'SC Local'                  : df.loc[i, "SC Local"],
                                    'IP Address'                : df.loc[i, "Net number / IP address"],
                                    'PLC No.'                   : df.loc[i, "PLC No"]
                                }
                    sum_list.append(sum_dict)
        return sum_list


    def CreateIPSheet(ipAdr_list): # ----------------Sheet for IP Addresses ----------------
 
        for i in range(len(df)):
            if str(df.loc[i, "Sort"]) == "1" and pd.notna(df.loc[i, "Component Description"]): area = str(df.loc[i, "Component Description"])              
            if pd.notna(df.loc[i, "Net number / IP address"]):
                    tagPart2 = df.loc[i, "TAG Part 2 (# Item)"]
                    tagPart2_noDot = tagPart2.split('.', 1)[0]
                    tagPart4 = df.loc[i, "TAG Part 4 (# Function serial)"]
                    tagPart4 = tagPart4.replace("TA", "MA")
                    identifier = str(df.loc[i, "Generator Function"])
                    if identifier == "FRQ_ETH": identifier = "FRQ"
                    ipAdr_dict = {'Tag Part 2'                  : df.loc[i, "TAG Part 2 (# Item)"],
                                    'Tag Part 3'                : df.loc[i, "TAG Part 3 (Function code)"],
                                    'Tag Part 4'                : df.loc[i, "TAG Part 4 (# Function serial)"],
                                    'Tag'                       : f'{tagPart2_noDot}_{df.loc[i, "TAG Part 3 (Function code)"]}_{tagPart4}',
                                    'Description'               : f'{df.loc[i, "Component Description"]}',
                                    'CM No'                     : df.loc[i, "CM No"],
                                    'IP-address'                : df.loc[i, "Net number / IP address"],
                                    'Area'                      : area,
                                    'Tag Name'                  : f'{tagPart2_noDot}_{df.loc[i, "TAG Part 3 (Function code)"]}_{tagPart4}_{identifier}_UDT',
                                    'PLC No.'                   : df.loc[i, "PLC No"],
                                    'Component Type'            : df.loc[i, "Component Type (Order number)"]
                                }
                    ipAdr_list.append(ipAdr_dict)
        return ipAdr_list


    #################################### Run Functions depending on which User Settings are Selected ####################################
    if Cmd_AssignCMNo:
        try: AssignCMNumber() 
        except Exception as e: 
            Sts_PrgStatus = f'Error while Assigning Control Module number: {e}'
            print(Sts_PrgStatus) 
    if Cmd_CreateSheets:
        try: CreateFRQSheet(frq_list)
        except Exception as e: 
            Sts_PrgStatus = f'Error while creating FRQ Sheet: {e}'
            print(Sts_PrgStatus)
        try: CreateMTRSheet(mtr_list)
        except Exception as e: 
            Sts_PrgStatus = f'Error while creating MTR Sheet: {e}'
            print(Sts_PrgStatus)
        try: CreateVLVSheet(vlv_list)
        except Exception as e: 
            Sts_PrgStatus = f'Error while creating VLV Sheet: {e}'
            print(Sts_PrgStatus)
        try: CreateDINSheet(din_list)
        except Exception as e: 
            Sts_PrgStatus = f'Error while creating DIN Sheet: {e}'
            print(Sts_PrgStatus)
        try: CreateAINSheet(ain_list)
        except Exception as e: 
            Sts_PrgStatus = f'Error while creating AIN Sheet: {e}'
            print(Sts_PrgStatus)
        try: CreateAOUTSheet(aout_list)
        except Exception as e: 
            Sts_PrgStatus = f'Error while creating AOUT Sheet: {e}'
            print(Sts_PrgStatus)
        try: CreatePIDSheet(pid_list)
        except Exception as e: 
            Sts_PrgStatus = f'Error while creating PID Sheet: {e}'
            print(Sts_PrgStatus)
        try: CreateALMSheet(alm_list)
        except Exception as e: 
            Sts_PrgStatus = f'Error while creating ALM Sheet: {e}'
            print(Sts_PrgStatus)
        try: CreateFPSheet(fp_list)
        except Exception as e: 
            Sts_PrgStatus = f'Error while creating FP Sheet: {e}'
            print(Sts_PrgStatus)
        try: CreateFPVLVSheet(fp_vlv_list)
        except Exception as e: 
            Sts_PrgStatus = f'Error while creating FP_VLV Sheet: {e}'
            print(Sts_PrgStatus)
        try: CreateSAFESheet(aoi_safe_list)
        except Exception as e: 
            Sts_PrgStatus = f'Error while creating SAFE Sheet: {e}'
            print(Sts_PrgStatus)
        try: CreateSUMSheet(sum_list)
        except Exception as e: 
            Sts_PrgStatus = f'Error while creating SUM Sheet: {e}'
            print(Sts_PrgStatus)
        try: CreateIPSheet(ipAdr_list)
        except Exception as e: 
            Sts_PrgStatus = f'Error while creating IP Sheet: {e}'
            print(Sts_PrgStatus)
        

    # Create dataframes to be exported to sheets
    frq_df = pd.DataFrame(frq_list) 
    mtr_df = pd.DataFrame(mtr_list)
    vlv_df = pd.DataFrame(vlv_list)
    din_df = pd.DataFrame(din_list)
    ain_df = pd.DataFrame(ain_list)
    aout_df = pd.DataFrame(aout_list)
    pid_df = pd.DataFrame(pid_list)
    alm_df = pd.DataFrame(alm_list)
    fp_df = pd.DataFrame(fp_list)
    fp_vlv_df = pd.DataFrame(fp_vlv_list)
    aoi_safe_df = pd.DataFrame(aoi_safe_list, dtype=object)
    sum_df = pd.DataFrame(sum_list)
    ipAdr_df = pd.DataFrame(ipAdr_list, dtype=object)

    # Sort rows in dataframes by CM No
    if Cmd_SortRows:
        try:
            if not frq_df.empty: frq_df = frq_df.sort_values(by ='CM No')
            if not mtr_df.empty: mtr_df = mtr_df.sort_values(by ='CM No')
            if not vlv_df.empty: vlv_df = vlv_df.sort_values(by ='CM No')
            if not din_df.empty: din_df = din_df.sort_values(by ='CM No')
            if not ain_df.empty: ain_df = ain_df.sort_values(by ='CM No')
            if not aout_df.empty: aout_df = aout_df.sort_values(by ='CM No')
            if not pid_df.empty: pid_df = pid_df.sort_values(by ='CM No')
            if not alm_df.empty: alm_df = alm_df.sort_values(by ='CM No')
            if not alm_df.empty: alm_df = alm_df.replace("N/A", "MSG", regex=True)
            if not fp_df.empty: fp_df = fp_df.sort_values(by ='CM No')
            if not fp_vlv_df.empty: fp_vlv_df = fp_vlv_df.sort_values(by ='CM No')
            if not aoi_safe_df.empty: aoi_safe_df = aoi_safe_df.sort_values(by ='CM No')
            if not sum_df.empty: sum_df = sum_df.sort_values(by ='CM No')
            if not ipAdr_df.empty: ipAdr_df = ipAdr_df.sort_values(by ='CM No')
        except Exception as e: 
            Sts_PrgStatus = f'Error while sorting rows: {e}'
            print(Sts_PrgStatus)


    # Check for duplicated CM No
    if Cmd_CheckDups:
        try:
            if not frq_df.empty: frq_df['Duplicates'] = np.where(frq_df['CM No'] == frq_df['CM No'].shift(), 'Duplicated', 'Not Duplicated')
            if not mtr_df.empty: mtr_df['Duplicates'] = np.where(mtr_df['CM No'] == mtr_df['CM No'].shift(), 'Duplicated', 'Not Duplicated')
            if not vlv_df.empty: vlv_df['Duplicates'] = np.where(vlv_df['CM No'] == vlv_df['CM No'].shift(), 'Duplicated', 'Not Duplicated')
            if not din_df.empty: din_df['Duplicates'] = np.where(din_df['CM No'] == din_df['CM No'].shift(), 'Duplicated', 'Not Duplicated')
            if not ain_df.empty: ain_df['Duplicates'] = np.where(ain_df['CM No'] == ain_df['CM No'].shift(), 'Duplicated', 'Not Duplicated')
            if not aout_df.empty: aout_df['Duplicates'] = np.where(aout_df['CM No'] == aout_df['CM No'].shift(), 'Duplicated', 'Not Duplicated')
            if not pid_df.empty: pid_df['Duplicates'] = np.where(pid_df['CM No'] == pid_df['CM No'].shift(), 'Duplicated', 'Not Duplicated')
            if not alm_df.empty: alm_df['Duplicates'] = np.where(alm_df['CM No'] == alm_df['CM No'].shift(), 'Duplicated', 'Not Duplicated')
            if not aoi_safe_df.empty: aoi_safe_df['Duplicates'] = np.where(aoi_safe_df['CM No'] == aoi_safe_df['CM No'].shift(), 'Duplicated', 'Not Duplicated')
            if not sum_df.empty: sum_df['Duplicates'] = np.where(sum_df['CM No'] == sum_df['CM No'].shift(), 'Duplicated', 'Not Duplicated')
            if not ipAdr_df.empty: ipAdr_df['Duplicates'] = np.where(ipAdr_df['CM No'] == ipAdr_df['CM No'].shift(), 'Duplicated', 'Not Duplicated')
        except Exception as e: 
            Sts_PrgStatus = f'Error while checking duplicated control module numbers: {e}'
            print(Sts_PrgStatus)


# ----------------Create all sheets and make column width dynamic----------------
    filename = "export/worksheet/AB_Worksheet.xlsx"

    def get_col_widths(dataframe):
        return [max([len(str(s)) for s in dataframe[col].values] + [len(col)]) for col in dataframe.columns]
    
    # If a cell is empty, Pandas interprets the cell as a float. To avoid this, fill empty cells.
    df.replace(np.nan, "", inplace=True)
    frq_df.replace(np.nan, "", inplace=True)
    mtr_df.replace(np.nan, "", inplace=True)
    vlv_df.replace(np.nan, "", inplace=True)
    din_df.replace(np.nan, "", inplace=True)
    ain_df.replace(np.nan, "", inplace=True)
    aout_df.replace(np.nan, "", inplace=True)
    pid_df.replace(np.nan, "", inplace=True)
    alm_df.replace(np.nan, "", inplace=True)
    fp_df.replace(np.nan, "", inplace=True)
    fp_vlv_df.replace(np.nan, "", inplace=True)
    aoi_safe_df.replace(np.nan, "", inplace=True)
    sum_df.replace(np.nan, "", inplace=True)
    ipAdr_df.replace(np.nan, "", inplace=True)

    # Create sheets
    if Cmd_CreateSheets:
        with pd.ExcelWriter(filename, engine='xlsxwriter') as writer:
            df.to_excel(writer, sheet_name='IO List', index=False)
            frq_df.to_excel(writer, sheet_name='FRQ', index=False)
            mtr_df.to_excel(writer, sheet_name='MTR', index=False)
            vlv_df.to_excel(writer, sheet_name='VLV', index=False)
            din_df.to_excel(writer, sheet_name='DIN', index=False)
            ain_df.to_excel(writer, sheet_name='AIN', index=False)
            aout_df.to_excel(writer, sheet_name='AOUT', index=False)
            pid_df.to_excel(writer, sheet_name='PID', index=False)
            alm_df.to_excel(writer, sheet_name='ALM', index=False)
            fp_df.to_excel(writer, sheet_name='FP', index=False)
            fp_vlv_df.to_excel(writer, sheet_name='FP_VLV', index=False)
            aoi_safe_df.to_excel(writer, sheet_name='SAFE', index=False)
            sum_df.to_excel(writer, sheet_name='SUM', index=False)
            ipAdr_df.to_excel(writer, sheet_name='IP Address', index=False)
            for i, width in enumerate(get_col_widths(df)):
                    writer.sheets['IO List'].set_column(i, i, width)    
            for i, width in enumerate(get_col_widths(frq_df)):
                    writer.sheets['FRQ'].set_column(i, i, width)
            for i, width in enumerate(get_col_widths(mtr_df)):
                    writer.sheets['MTR'].set_column(i, i, width)
            for i, width in enumerate(get_col_widths(vlv_df)):
                    writer.sheets['VLV'].set_column(i, i, width)
            for i, width in enumerate(get_col_widths(din_df)):
                    writer.sheets['DIN'].set_column(i, i, width)
            for i, width in enumerate(get_col_widths(ain_df)):
                    writer.sheets['AIN'].set_column(i, i, width)
            for i, width in enumerate(get_col_widths(aout_df)):
                    writer.sheets['AOUT'].set_column(i, i, width)
            for i, width in enumerate(get_col_widths(pid_df)):
                    writer.sheets['PID'].set_column(i, i, width)
            for i, width in enumerate(get_col_widths(alm_df)):
                    writer.sheets['ALM'].set_column(i, i, width)
            for i, width in enumerate(get_col_widths(fp_df)):
                    writer.sheets['FP'].set_column(i, i, width)
            for i, width in enumerate(get_col_widths(fp_vlv_df)):
                    writer.sheets['FP_VLV'].set_column(i, i, width)
            for i, width in enumerate(get_col_widths(aoi_safe_df)):
                    writer.sheets['SAFE'].set_column(i, i, width)
            for i, width in enumerate(get_col_widths(sum_df)):
                    writer.sheets['SUM'].set_column(i, i, width)
            for i, width in enumerate(get_col_widths(ipAdr_df)):
                    writer.sheets['IP Address'].set_column(i, i, width)

    Sts_PrgStatus = "Completed"


# ----------------Styles for worksheet----------------

def StyleWS():
    filepath = r'export/worksheet/AB_Worksheet.xlsx'
    wb = load_workbook(filepath)

    bg_color = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    border_color = Border(right=Side(style='thin'))
    font_color = Font(color='000000', bold=True)


# ----------------Apply styles to sheets----------------
    ws = wb['IO List']
    row_count = ws.max_row + 1
    for x in range(row_count):
        if x % 2 != 0:
            for cell in ws[x:x]:
                cell.fill = bg_color
    
    ws = wb['FRQ']
    row_count = ws.max_row + 1
    for x in range(row_count):
        if x % 2 != 0:
            for cell in ws[x:x]:
                cell.fill = bg_color
    
    ws = wb['MTR']
    row_count = ws.max_row + 1
    for x in range(row_count):
        if x % 2 != 0:
            for cell in ws[x:x]:
                cell.fill = bg_color
    
    ws = wb['VLV']
    row_count = ws.max_row + 1
    for x in range(row_count):
        if x % 2 != 0:
            for cell in ws[x:x]:
                cell.fill = bg_color
    
    ws = wb['DIN']
    row_count = ws.max_row + 1
    for x in range(row_count):
        if x % 2 != 0:
            for cell in ws[x:x]:
                cell.fill = bg_color
    
    ws = wb['AIN']
    row_count = ws.max_row + 1
    for x in range(row_count):
        if x % 2 != 0:
            for cell in ws[x:x]:
                cell.fill = bg_color
    
    ws = wb['AOUT']
    row_count = ws.max_row + 1
    for x in range(row_count):
        if x % 2 != 0:
            for cell in ws[x:x]:
                cell.fill = bg_color
    
    ws = wb['PID']
    row_count = ws.max_row + 1
    for x in range(row_count):
        if x % 2 != 0:
            for cell in ws[x:x]:
                cell.fill = bg_color
    
    ws = wb['ALM']
    row_count = ws.max_row + 1
    for x in range(row_count):
        if x % 2 != 0:
            for cell in ws[x:x]:
                cell.fill = bg_color
    
    ws = wb['FP']
    row_count = ws.max_row + 1
    for x in range(row_count):
        if x % 2 != 0:
            for cell in ws[x:x]:
                cell.fill = bg_color
    
    ws = wb['FP_VLV']
    row_count = ws.max_row + 1
    for x in range(row_count):
        if x % 2 != 0:
            for cell in ws[x:x]:
                cell.fill = bg_color
    
    ws = wb['SAFE']
    row_count = ws.max_row + 1
    for x in range(row_count):
        if x % 2 != 0:
            for cell in ws[x:x]:
                cell.fill = bg_color
    
    ws = wb['SUM']
    row_count = ws.max_row + 1
    for x in range(row_count):
        if x % 2 != 0:
            for cell in ws[x:x]:
                cell.fill = bg_color
    
    ws = wb['IP Address']
    row_count = ws.max_row + 1
    for x in range(row_count):
        if x % 2 != 0:
            for cell in ws[x:x]:
                cell.fill = bg_color
    
    wb.save(filepath)
    wb.close()