import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import colors, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def ExportWS(df, Cmd_SortRows, Cmd_CheckDups, Cmd_AssignCMNo, Cmd_CreateSheets):

    # Initialization of Dataframe   
    df.columns = df.iloc[2].astype(str) # Set the header to be at index 2
    df = pd.DataFrame(df.iloc[3:, 1:]) # Remove the first 3 rows and the first column
    df = df.reset_index()

    # Creates lists for dataframes
    frq_list = list()
    mtr_list = list()
    vlv_list = list()
    din_list = list()
    ain_list = list()
    aout_list = list()
    pid_list = list()
    alm_list = list()
    fp_list = list()
    fp_vlv_list = list()
    aoi_safe_list = list()
    sum_list = list()
    ipAdr_list = list()
    hw_list = list()
    area_list = list()
    cable_list = list()

    # Status variable to display current program state
    global Sts_PrgStatus

    identifiers = ["FRQ", "MTR", "AIN", "AOUT", "DIN", "PID", "ALM", "VLV"]

    Sts_PrgStatus = "Please change Worksheet settings"

    def AssignCMNumber(): # ----------------Autogenerate Generator Config Number----------------
        ##### FRQ #####
        num_count = 0
        next_num = 0
        num_count_list = []
        for i in range(len(df)): # Check is CM number is already assigned and get the highest value            
            identifier = df.loc[i,"CM Type"]            
            if identifier == "FRQ" and pd.notna(df.loc[i, "CM No"]):
                num_count = df.loc[i, "CM No"]
                num_count_list.append(num_count)
        

        if num_count_list == []: # Assign starting CM number
            next_num = 1000
        else:
            next_num = max(num_count_list)        

        for i in range(len(df)):                      
            identifier = df.loc[i,"CM Type"]                       
            tag_part_3 = df.loc[i, "TAG Part 3 (Function code)"]                        
            if identifier == "FRQ" and pd.isna(df.loc[i, "CM No"]):
                next_num = next_num + 1
                df.loc[i, "CM No"] = next_num
                for x in df.index:
                    identifier = df.loc[x, "CM Type"]                    
                    if (identifier == "FRQ_Ptc") and (df.loc[x, "TAG Part 3 (Function code)"] == tag_part_3):
                        df.loc[x, "CM No"] = next_num
                    elif (identifier == "FRQ_Pro") and (df.loc[x, "TAG Part 3 (Function code)"] == tag_part_3):
                        df.loc[x, "CM No"] = next_num
                    elif (identifier == "FRQ_R") and (df.loc[x, "TAG Part 3 (Function code)"] == tag_part_3):
                        df.loc[x, "CM No"] = next_num
                    elif (identifier == "FRQ_Dis") and (df.loc[x, "TAG Part 3 (Function code)"] == tag_part_3):
                        df.loc[x, "CM No"] = next_num
                    elif (identifier == "FRQ_Rot") and (df.loc[x, "TAG Part 3 (Function code)"] == tag_part_3):
                        df.loc[x, "CM No"] = next_num
                    elif (identifier == "FRQ_IO") and (df.loc[x, "TAG Part 3 (Function code)"] == tag_part_3):
                        df.loc[x, "CM No"] = next_num
        

        # ##### MTR #####
        num_count = 0
        next_num = 0
        num_count_list = []
        for i in range(len(df)): # Check is CM number is already assigned and get the highest value
            identifier = df.loc[i,"CM Type"]
            if identifier == "MTR" and pd.notna(df.loc[i, "CM No"]):
                num_count = df.loc[i, "CM No"]
                num_count_list.append(num_count)

        if num_count_list == []: # Assign starting CM number
            next_num = 2000
        else:
            next_num = max(num_count_list)

        for i in range(len(df)):
            identifier = df.loc[i, "CM Type"]
            tag_part_3 = df.loc[i, "TAG Part 3 (Function code)"]
            if identifier == "MTR" and pd.isna(df.loc[i, "CM No"]):
                next_num = next_num + 1
                df.loc[i, "CM No"] = next_num
                for x in range(len(df)):
                    identifier = df.loc[x, "CM Type"]
                    if (identifier == "MTR_Dis") and (df.loc[x, "TAG Part 3 (Function code)"] == tag_part_3):
                        df.loc[x, "CM No"] = next_num
                    elif (identifier == "MTR_Pro") and (df.loc[x, "TAG Part 3 (Function code)"] == tag_part_3):
                        df.loc[x, "CM No"] = next_num
                    elif (identifier == "MTR_Out") and (df.loc[x, "TAG Part 3 (Function code)"] == tag_part_3):
                        df.loc[x, "CM No"] = next_num
                    elif (identifier == "MTR_Run") and (df.loc[x, "TAG Part 3 (Function code)"] == tag_part_3):
                        df.loc[x, "CM No"] = next_num
                    elif (identifier == "MTR_Rot") and (df.loc[x, "TAG Part 3 (Function code)"] == tag_part_3):
                        df.loc[x, "CM No"] = next_num
                    elif (identifier == "MTR_Saf") and (df.loc[x, "TAG Part 3 (Function code)"] == tag_part_3):
                        df.loc[x, "CM No"] = next_num
                    elif (identifier == "MTR_Saf_Fb") and (df.loc[x, "TAG Part 3 (Function code)"] == tag_part_3):
                        df.loc[x, "CM No"] = next_num


        ##### AIN #####
        num_count = 0
        next_num = 0
        num_count_list = []
        for i in range(len(df)): # Check is CM number is already assigned and get the highest value
            identifier = df.loc[i,"CM Type"]
            if identifier == "AIN" and pd.notna(df.loc[i, "CM No"]):
                num_count = df.loc[i, "CM No"]
                num_count_list.append(num_count)

        if num_count_list == []: # Assign starting CM number
            next_num = 3000
        else:
            next_num = max(num_count_list)

        for i in range(len(df)):
            identifier = df.loc[i,"CM Type"]
            if identifier == "AIN" and pd.isna(df.loc[i, "CM No"]):
                next_num = next_num + 1
                df.loc[i, "CM No"] = next_num


        ##### AOUT #####
        num_count = 0
        next_num = 0
        num_count_list = []
        for i in range(len(df)): # Check is CM number is already assigned and get the highest value
            identifier = df.loc[i,"CM Type"]
            if identifier == "AOUT" and pd.notna(df.loc[i, "CM No"]):
                num_count = df.loc[i, "CM No"]
                num_count_list.append(num_count)

        if num_count_list == []: # Assign starting CM number
            next_num = 4000
        else:
            next_num = max(num_count_list)

        for i in range(len(df)):
            identifier = df.loc[i,"CM Type"]
            if identifier == "AOUT" and pd.isna(df.loc[i, "CM No"]):
                next_num = next_num + 1
                df.loc[i, "CM No"] = next_num


        ##### DIN #####
        num_count = 0
        next_num = 0
        num_count_list = []
        for i in range(len(df)): # Check is CM number is already assigned and get the highest value
            identifier = df.loc[i,"CM Type"]
            if identifier == "DIN" and pd.notna(df.loc[i, "CM No"]):
                num_count = df.loc[i, "CM No"]
                num_count_list.append(num_count)

        if num_count_list == []: # Assign starting CM number
            next_num = 5000
        else:
            next_num = max(num_count_list)

        for i in range(len(df)):
            identifier = df.loc[i,"CM Type"]
            if identifier == "DIN" and pd.isna(df.loc[i, "CM No"]):
                next_num = next_num + 1
                df.loc[i, "CM No"] = next_num


        ##### PID #####
        num_count = 0
        next_num = 0
        num_count_list = []
        for i in range(len(df)): # Check is CM number is already assigned and get the highest value
            identifier = df.loc[i,"CM Type"]
            if identifier == "PID" and pd.notna(df.loc[i, "CM No"]):
                num_count = df.loc[i, "CM No"]
                num_count_list.append(num_count)

        if num_count_list == []: # Assign starting CM number
            next_num = 6000
        else:
            next_num = max(num_count_list)

        for i in range(len(df)):
            identifier = df.loc[i,"CM Type"]
            if identifier == "PID" and pd.isna(df.loc[i, "CM No"]):
                next_num = next_num + 1
                df.loc[i, "CM No"] = next_num


        ##### ALM #####
        num_count = 0
        next_num = 0
        num_count_list = []
        for i in range(len(df)): # Check is CM number is already assigned and get the highest value
            identifier = df.loc[i,"CM Type"]
            if identifier == "ALM" and pd.notna(df.loc[i, "CM No"]):
                num_count = df.loc[i, "CM No"]
                num_count_list.append(num_count)

        if num_count_list == []: # Assign starting CM number
            next_num = 10000
        else:
            next_num = max(num_count_list)
        
        for i in range(len(df)):
            identifier = df.loc[i,"CM Type"]
            if identifier == "ALM" and pd.isna(df.loc[i, "CM No"]):
                next_num = next_num + 1
                df.loc[i, "CM No"] = next_num


        ##### VLV #####
        num_count = 0
        next_num = 0
        num_count_list = []
        for i in range(len(df)): # Check is CM number is already assigned and get the highest value
            identifier = df.loc[i,"CM Type"]
            if identifier == "VLV" and pd.notna(df.loc[i, "CM No"]):
                num_count = df.loc[i, "CM No"]
                num_count_list.append(num_count)

        if num_count_list == []: # Assign starting CM number
            next_num = 7000
        else:
            next_num = max(num_count_list)

        for i in range(len(df)):
            identifier = df.loc[i,"CM Type"]
            tag_part_3 = df.loc[i, "TAG Part 3 (Function code)"]
            if identifier == "VLV" and pd.isna(df.loc[i, "CM No"]):
                next_num = next_num + 1
                df.loc[i, "CM No"] = next_num
                for x in range(len(df)):
                    identifier = df.loc[x,"CM Type"]
                    if (identifier == "VLV_N") and (df.loc[i, "TAG Part 3 (Function code)"] == tag_part_3):
                        df.loc[i, "CM No"] = next_num
                    elif (identifier == "VLV_OutN") and (df.loc[i, "TAG Part 3 (Function code)"] == tag_part_3):
                        df.loc[i, "CM No"] = next_num
                    elif (identifier == "VLV_OutP") and (df.loc[i, "TAG Part 3 (Function code)"] == tag_part_3):
                        df.loc[i, "CM No"] = next_num
                    elif (identifier == "VLV_P") and (df.loc[i, "TAG Part 3 (Function code)"] == tag_part_3):
                        df.loc[i, "CM No"] = next_num


        ##### SUM #####
        num_count = 0
        next_num = 0
        num_count_list = []
        for i in range(len(df)): # Check is CM number is already assigned and get the highest value
            identifier = df.loc[i,"CM Type"]
            if identifier == "SUM" and pd.notna(df.loc[i, "CM No"]):
                num_count = df.loc[i, "CM No"]
                num_count_list.append(num_count)

        if num_count_list == []:
            next_num = 8000
        else:
            next_num = max(num_count_list)

        for i in range(len(df)):
            identifier = df.loc[i,"CM Type"]
            if identifier == "SUM" and pd.isna(df.loc[i, "CM No"]):
                next_num = next_num + 1
                df.loc[i, "CM No"] = next_num
    
    def CreateFRQSheet(frq_list): # ----------------Sheet for FRQ component----------------        
        
        for i in range(len(df)):
            # if str(df.loc[i, "Sort"]) == "1" and pd.notna(df.loc[i, "Component Description"]): area = df.loc[i, "Component Description"]
            identifier = df.loc[i,"CM Type"]           
            if identifier == ("FRQ"):
                tagPart2 = df.loc[i, "TAG Part 2 (# Item)"]                
                
                tagPart2_noDot = tagPart2.split('.', 1)[0]
                frq_dict = {'Tag Part 1'                : df.loc[i, "TAG Part 1 (# System)"],
                            'Tag Part 2'                : df.loc[i, "TAG Part 2 (# Item)"],
                            'Tag Part 3'                : df.loc[i, "TAG Part 3 (Function code)"],
                            'Tag Part 4'                : df.loc[i, "TAG Part 4 (# Function serial)"],
                            'Tag'                       : f'{tagPart2_noDot}_{df.loc[i, "TAG Part 3 (Function code)"]}_{df.loc[i, "TAG Part 4 (# Function serial)"]}',
                            'Description'               : f'{df.loc[i, "Component Description"]} - {df.loc[i, "Process Description"]}',
                            'CM Type'                   : df.loc[i,"CM Type"],
                            'CM No'                     : int(df.loc[i, "CM No"]),
                            'Cabinet'                   : df.loc[i, "TAG Part 2 (# Item)"],
                            'Area'                      : f'{df.loc[i, "Tab ID"]}{df.loc[i, "Sub ID"]}',
                            'PLC No.'                   : df.loc[i, "PLC NO"],
                            'Tab ID'                    : df.loc[i, "Tab ID"],
                            'Sub ID'                    : df.loc[i, "Sub ID"],
                            'SC_Local'                  : df.loc[i, "SC Local"],
                            'FRQ Type'                  : '',
                            'Thermal Fault'             : '',
                            'Rotation Sensor'           : '',
                            'Thermistor'                : '',
                            'Disconnect'                : '',
                            'Safety'                    : df.loc[i, "Safety - F_Dest_Add (HEX)"],
                            'Main Power [KW]'           : df.loc[i, "Main Power [kW]"],
                            'Main Current [A]'          : df.loc[i, "Main Current [A]"],
                            'Main Voltage [V]'          : df.loc[i, "Main Voltage [V]"],
                            'Connection'                : df.loc[i, "Motor Connection"],
                            'Poles'                     : df.loc[i, "Poles"],
                            'Frequency [Hz]'            : df.loc[i, "Frequency [Hz]"],
                            'Cos φ'                     : df.loc[i, "Cos φ (Efficiency)"],
                            'RPM'                       : df.loc[i, "RPM"],
                            'EU'                        : df.loc[i, "EU"],
                            'Min EU'                    : df.loc[i, "Min EU"],
                            'Max EU'                    : df.loc[i, "Max EU"],
                            'Raw Min'                   : df.loc[i, "Raw Min"],
                            'Raw Max'                   : df.loc[i, "Raw Max"],
                            'IP Address'                : '',
                            'Electrical Room'           : '',
                            'Pdf Elect Schema Page'     : df.loc[i, "Pdf Elect Schema Page"],
                            'Pdf Vfd Schema Page'       : df.loc[i, "Pdf Vfd Schema Page"],
                            'Checked'                   : ''
                            }
                frq_list.append(frq_dict)

    # Search the list for a dict with a matching generator config, and modify the corresponding dict.
        for i in range(len(df)):
            identifier = df.loc[i,"CM Type"]
            if identifier == ("FRQ_Pro"):
                gen_con = df.loc[i, "CM No"]
                list_index = next((index for (index, d) in enumerate(frq_list) if d["CM No"] == gen_con), None)
                if list_index is not None:
                    frq_list[int(list_index)]['Thermal Fault'] = df.loc[i, "PLC Tag Name"]
        for i in range(len(df)):
            identifier = df.loc[i,"CM Type"]
            if identifier == ("FRQ_Rot"):
                gen_con = df.loc[i, "CM No"]
                list_index = next((index for (index, d) in enumerate(frq_list) if d["CM No"] == gen_con), None)
                if list_index is not None:
                    frq_list[int(list_index)]['Rotation Sensor'] = df.loc[i, "PLC Tag Name"]
        for i in range(len(df)):
            identifier = df.loc[i,"CM Type"]
            if identifier == ("FRQ_Ptc"):
                gen_con = df.loc[i, "CM No"]
                list_index = next((index for (index, d) in enumerate(frq_list) if d["CM No"] == gen_con), None)
                if list_index is not None:
                    frq_list[int(list_index)]['Thermistor'] = df.loc[i, "Process Description"]
        for i in range(len(df)):
            identifier = df.loc[i,"CM Type"]
            if identifier == ("FRQ_Dis"):
                gen_con = df.loc[i, "CM No"]
                list_index = next((index for (index, d) in enumerate(frq_list) if d["CM No"] == gen_con), None)
                if list_index is not None:
                    frq_list[int(list_index)]['Disconnect'] = df.loc[i, "PLC Tag Name"]
        for i in range(len(df)):
            identifier = df.loc[i,"CM Type"]
            if identifier == ("FRQ_IO"):
                gen_con = df.loc[i, "CM No"]
                list_index = next((index for (index, d) in enumerate(frq_list) if d["CM No"] == gen_con), None)
                if list_index is not None:
                    frq_list[int(list_index)]['IP Address'] = df.loc[i, "Net number / IP address"]
                    frq_list[int(list_index)]['FRQ Type'] = df.loc[i, "FRQ Type"]
                    frq_list[int(list_index)]['Electrical Room'] = df.loc[i, "TAG Part 1 (# System)"]                
        return frq_list
    
    def CreateMTRSheet(mtr_list): # ----------------Sheet for MTR component----------------        

        for i in range(len(df)):
            # if str(df.loc[i, "Sort"]) == "1" and pd.notna(df.loc[i, "Component Description"]): area = df.loc[i, "Component Description"]
            identifier = df.loc[i,"CM Type"]
            if identifier == ("MTR"):
                tagPart2 = df.loc[i, "TAG Part 2 (# Item)"]
                tagPart2_noDot = tagPart2.split('.', 1)[0]
                mtr_dict = {'Tag Part 1'                : df.loc[i, "TAG Part 1 (# System)"],
                            'Tag Part 2'                : df.loc[i, "TAG Part 2 (# Item)"],
                            'Tag Part 3'                : df.loc[i, "TAG Part 3 (Function code)"],
                            'Tag Part 4'                : df.loc[i, "TAG Part 4 (# Function serial)"],
                            'Tag'                       : f'{tagPart2_noDot}_{df.loc[i, "TAG Part 3 (Function code)"]}_{df.loc[i, "TAG Part 4 (# Function serial)"]}',
                            'Description'               : f'{df.loc[i, "Component Description"]} - {df.loc[i, "Process Description"]}',
                            'CM Type'                   : df.loc[i,"CM Type"],
                            'CM No'                     : int(df.loc[i, "CM No"]),
                            'Cabinet'                   : df.loc[i, "TAG Part 2 (# Item)"],
                            'Area'                      : f'{df.loc[i, "Tab ID"]}{df.loc[i, "Sub ID"]}',
                            'PLC No.'                   : df.loc[i, "PLC NO"],
                            'Tab ID'                    : df.loc[i, "Tab ID"],
                            'Sub ID'                    : df.loc[i, "Sub ID"],
                            'SC_Local'                  : df.loc[i, "SC Local"],
                            'Thermal Fault'                 : '',
                            'Rotation Sensor'               : '',
                            'Thermistor'                    : '',
                            'Start Forward'                 : '',
                            'Start Reverse'                 : '',
                            'Forward Feedback'              : '',
                            'Reverse Feedback'              : '',
                            'Emergency Stop Contactor'      : '',
                            'Emergency Stop Contactor FB'   : '',
                            'Disconnect'                    : '',
                            'Forward PLC address'           : '',
                            'Electrical Room'               : '',
                            'Pdf Elect Schema Page'         : df.loc[i, "Pdf Elect Schema Page"],
                            'Pdf Vfd Schema Page'           : df.loc[i, "Pdf Vfd Schema Page"],
                            'Checked'                       : ''
                            }
                mtr_list.append(mtr_dict)

        # Search the list for a dict with a matching generator config, and modify the corresponding dict.
        for i in range(len(df)):
            identifier = df.loc[i,"CM Type"]
            if identifier == ("MTR_Dis"):
                gen_con = df.loc[i, "CM No"]
                list_index = next((index for (index, d) in enumerate(mtr_list) if d["CM No"] == gen_con), None)
                if list_index is not None:
                    mtr_list[int(list_index)]['Disconnect'] = df.loc[i, "PLC Tag Name"]
        for i in range(len(df)):
            identifier = df.loc[i,"CM Type"]
            if identifier == ("MTR_Pro"):
                gen_con = df.loc[i, "CM No"]
                list_index = next((index for (index, d) in enumerate(mtr_list) if d["CM No"] == gen_con), None)
                if list_index is not None:
                    mtr_list[int(list_index)]['Thermal Fault'] = df.loc[i, "PLC Tag Name"]
                    mtr_list[int(list_index)]['Electrical Room'] = df.loc[i, "TAG Part 1 (# System)"]
        for i in range(len(df)):
            identifier = df.loc[i,"CM Type"]
            if identifier == ("MTR_Rot"):
                gen_con = df.loc[i, "CM No"]
                list_index = next((index for (index, d) in enumerate(mtr_list) if d["CM No"] == gen_con), None)
                if list_index is not None:
                    mtr_list[int(list_index)]['Rotation Sensor'] = df.loc[i, "PLC Tag Name"]
        for i in range(len(df)):
            identifier = df.loc[i,"CM Type"]
            if identifier == ("MTR_Out"):
                gen_con = df.loc[i, "CM No"]
                list_index = next((index for (index, d) in enumerate(mtr_list) if d["CM No"] == gen_con), None)
                if list_index is not None:
                    mtr_list[int(list_index)]['Start Forward'] = df.loc[i, "PLC Tag Name"]
                    mtr_list[int(list_index)]['Forward PLC address'] = df.loc[i, "PLC Address"]
        for i in range(len(df)):
            identifier = df.loc[i,"CM Type"]
            if identifier == ("MTR_Out_Rev"):
                gen_con = df.loc[i, "CM No"]
                list_index = next((index for (index, d) in enumerate(mtr_list) if d["CM No"] == gen_con), None)
                if list_index is not None:
                    mtr_list[int(list_index)]['Start Reverse'] = df.loc[i, "PLC Tag Name"]
        for i in range(len(df)):
            identifier = df.loc[i,"CM Type"]
            if identifier == ("MTR_Run"):
                gen_con = df.loc[i, "CM No"]
                list_index = next((index for (index, d) in enumerate(mtr_list) if d["CM No"] == gen_con), None)
                if list_index is not None:
                    mtr_list[int(list_index)]['Forward Feedback'] = df.loc[i, "PLC Tag Name"]
        for i in range(len(df)):
            identifier = df.loc[i,"CM Type"]
            if identifier == ("MTR_Run_Rev"):
                gen_con = df.loc[i, "CM No"]
                list_index = next((index for (index, d) in enumerate(mtr_list) if d["CM No"] == gen_con), None)
                if list_index is not None:
                    mtr_list[int(list_index)]['Reverse Feedback'] = df.loc[i, "PLC Tag Name"]
        for i in range(len(df)):
            identifier = df.loc[i,"CM Type"]
            if identifier == ("MTR_Saf"):
                gen_con = df.loc[i, "CM No"]
                list_index = next((index for (index, d) in enumerate(mtr_list) if d["CM No"] == gen_con), None)
                if list_index is not None:
                    mtr_list[int(list_index)]['Emergency Stop Contactor'] = df.loc[i, "PLC Tag Name"]
        for i in range(len(df)):
            identifier = df.loc[i,"CM Type"]
            if identifier == ("MTR_Saf_Fb"):
                gen_con = df.loc[i, "CM No"]
                list_index = next((index for (index, d) in enumerate(mtr_list) if d["CM No"] == gen_con), None)
                if list_index is not None:
                    mtr_list[int(list_index)]['Emergency Stop Contactor FB'] = df.loc[i, "PLC Tag Name"]

        return mtr_list

    def CreateVLVSheet(vlv_list): # ----------------Sheet for VLV component----------------
        for i in range(len(df)):
            # if str(df.loc[i, "Sort"]) == "1" and pd.notna(df.loc[i, "Component Description"]): area = df.loc[i, "Component Description"]
            identifier = df.loc[i,"CM Type"]
            if identifier == ("VLV"):
                tagPart2 = df.loc[i, "TAG Part 2 (# Item)"]
                tagPart2_noDot = tagPart2.split('.', 1)[0]
                vlv_dict = {'Tag Part 1'                : df.loc[i, "TAG Part 1 (# System)"],
                            'Tag Part 2'                : df.loc[i, "TAG Part 2 (# Item)"],
                            'Tag Part 3'                : df.loc[i, "TAG Part 3 (Function code)"],
                            'Tag Part 4'                : df.loc[i, "TAG Part 4 (# Function serial)"],
                            'Tag'                       : f'{tagPart2_noDot}_{df.loc[i, "TAG Part 3 (Function code)"]}_{df.loc[i, "TAG Part 4 (# Function serial)"]}',
                            'Description'               : f'{df.loc[i, "Component Description"]} - {df.loc[i, "Process Description"]}',
                            'CM Type'                   : df.loc[i,"CM Type"],
                            'CM No'                     : int(df.loc[i, "CM No"]),
                            'Cabinet'                   : df.loc[i, "TAG Part 2 (# Item)"],
                            'Area'                      : f'{df.loc[i, "Tab ID"]}{df.loc[i, "Sub ID"]}',
                            'PLC No.'                   : df.loc[i, "PLC NO"],
                            'Tab ID'                    : df.loc[i, "Tab ID"],
                            'Sub ID'                    : df.loc[i, "Sub ID"],
                            'SC_Local'                  : df.loc[i, "SC Local"],
                            'Cfg Type'                  : df.loc[i, "Cfg Type"],
                            'Valve Out P'               : '',
                            'Valve Out N'               : '',
                            'Valve FB P'                : '',
                            'Valve FB N'                : '',
                            'Activate Text'             : df.loc[i, "Activate Text"],
                            'Deactivate Text'           : df.loc[i, "Deactivate Text"],
                            'Pdf Elect Schema Page'     : df.loc[i, "Pdf Elect Schema Page"],
                            'Pdf Vfd Schema Page'       : df.loc[i, "Pdf Vfd Schema Page"],
                            'Checked'                   : ''
                            }
                vlv_list.append(vlv_dict)

        for i in range(len(df)):
            identifier = df.loc[i,"CM Type"]
            if identifier == ("VLV_OutP"):
                gen_con = df.loc[i, "CM No"]
                list_index = next((index for (index, d) in enumerate(vlv_list) if d["CM No"] == gen_con), None)
                if list_index is not None:
                    vlv_list[int(list_index)]['Valve Out P'] = df.loc[i, "PLC Tag Name"]
                    vlv_list[int(list_index)]['IO-address'] = df.loc[i, "PLC Address"]
        for i in range(len(df)):
            identifier = df.loc[i,"CM Type"]
            if identifier == ("VLV_OutN"):
                gen_con = df.loc[i, "CM No"]
                list_index = next((index for (index, d) in enumerate(vlv_list) if d["CM No"] == gen_con), None)
                if list_index is not None:
                    vlv_list[int(list_index)]['Valve Out N'] = df.loc[i, "PLC Tag Name"]
        for i in range(len(df)):
            identifier = df.loc[i,"CM Type"]
            if identifier == ("VLV_P"):
                gen_con = df.loc[i, "CM No"]
                list_index = next((index for (index, d) in enumerate(vlv_list) if d["CM No"] == gen_con), None)
                if list_index is not None:
                    vlv_list[int(list_index)]['Valve FB P'] = df.loc[i, "PLC Tag Name"]
        for i in range(len(df)):
            identifier = df.loc[i,"CM Type"]
            if identifier == ("VLV_N"):
                gen_con = df.loc[i, "CM No"]
                list_index = next((index for (index, d) in enumerate(vlv_list) if d["CM No"] == gen_con), None)
                if list_index is not None:
                    vlv_list[int(list_index)]['Valve FB N'] = df.loc[i, "PLC Tag Name"]

    def CreateDINSheet(din_list): # ----------------Sheet for DIN component----------------
        for i in range(len(df)):
            # if str(df.loc[i, "Sort"]) == "1" and pd.notna(df.loc[i, "Component Description"]): area = df.loc[i, "Component Description"]
            identifier = df.loc[i,"CM Type"]
            if identifier == ("DIN"):
                tagPart2 = df.loc[i, "TAG Part 2 (# Item)"]
                tagPart2_noDot = tagPart2.split('.', 1)[0]
                din_dict = {'Tag Part 1'                : df.loc[i, "TAG Part 1 (# System)"],
                            'Tag Part 2'                : df.loc[i, "TAG Part 2 (# Item)"],
                            'Tag Part 3'                : df.loc[i, "TAG Part 3 (Function code)"],
                            'Tag Part 4'                : df.loc[i, "TAG Part 4 (# Function serial)"],
                            'Tag'                       : f'{tagPart2_noDot}_{df.loc[i, "TAG Part 3 (Function code)"]}_{df.loc[i, "TAG Part 4 (# Function serial)"]}',
                            'Description'               : f'{df.loc[i, "Component Description"]} - {df.loc[i, "Process Description"]}',
                            'CM Type'                   : df.loc[i,"CM Type"],
                            'CM No'                     : int(df.loc[i, "CM No"]),
                            'Cabinet'                   : df.loc[i, "TAG Part 2 (# Item)"],
                            'Area'                      : f'{df.loc[i, "Tab ID"]}{df.loc[i, "Sub ID"]}',
                            'PLC No.'                   : df.loc[i, "PLC NO"],
                            'Tab ID'                    : df.loc[i, "Tab ID"],
                            'Sub ID'                    : df.loc[i, "Sub ID"],
                            'SC_Local'                  : df.loc[i, "SC Local"],
                            'Cfg Type'                  : df.loc[i, "Cfg Type"],
                            'Cfg Invert'                : df.loc[i, "Cfg Invert"],
                            'PLC Tag Name'              : df.loc[i, "PLC Tag Name"],
                            'PLC Tag Address'           : df.loc[i, "PLC Address"],
                            'Pdf Elect Schema Page'     : df.loc[i, "Pdf Elect Schema Page"],
                            'Pdf Vfd Schema Page'       : df.loc[i, "Pdf Vfd Schema Page"],
                            'Checked'                   : ''
                            }
                din_list.append(din_dict)

        return din_list

    def CreateAINSheet(ain_list): # ----------------Sheet for AIN component----------------
        for i in range(len(df)):
            # if str(df.loc[i, "Sort"]) == "1" and pd.notna(df.loc[i, "Component Description"]): area = df.loc[i, "Component Description"]
            identifier = df.loc[i,"CM Type"]
            if identifier == ("AIN"):
                tagPart2 = df.loc[i, "TAG Part 2 (# Item)"]
                tagPart2_noDot = tagPart2.split('.', 1)[0]
                ain_dict = {'Tag Part 1'                : df.loc[i, "TAG Part 1 (# System)"],
                            'Tag Part 2'                : df.loc[i, "TAG Part 2 (# Item)"],
                            'Tag Part 3'                : df.loc[i, "TAG Part 3 (Function code)"],
                            'Tag Part 4'                : df.loc[i, "TAG Part 4 (# Function serial)"],
                            'Tag'                       : f'{tagPart2_noDot}_{df.loc[i, "TAG Part 3 (Function code)"]}_{df.loc[i, "TAG Part 4 (# Function serial)"]}',
                            'Description'               : f'{df.loc[i, "Component Description"]} - {df.loc[i, "Process Description"]}',
                            'CM Type'                   : df.loc[i,"CM Type"],
                            'CM No'                     : int(df.loc[i, "CM No"]),
                            'Cabinet'                   : df.loc[i, "TAG Part 2 (# Item)"],
                            'Area'                      : f'{df.loc[i, "Tab ID"]}{df.loc[i, "Sub ID"]}',
                            'PLC No.'                   : df.loc[i, "PLC NO"],
                            'Tab ID'                    : df.loc[i, "Tab ID"],
                            'Sub ID'                    : df.loc[i, "Sub ID"],
                            'SC_Local'                  : df.loc[i, "SC Local"],
                            'PLC Tag Name'              : df.loc[i, "PLC Tag Name"],
                            'PLC Tag Address'           : df.loc[i, "PLC Address"],
                            'EU'                        : df.loc[i, "EU"],
                            'Min EU'                    : df.loc[i, "Min EU"],
                            'Max EU'                    : df.loc[i, "Max EU"],
                            'Raw Min'                   : df.loc[i, "Raw Min"],
                            'Raw Max'                   : df.loc[i, "Raw Max"],
                            'Pdf Elect Schema Page'     : df.loc[i, "Pdf Elect Schema Page"],
                            'Pdf Vfd Schema Page'       : df.loc[i, "Pdf Vfd Schema Page"],
                            'Checked'                   : ''
                            }
                ain_list.append(ain_dict)

        return ain_list

    def CreateAOUTSheet(aout_list): # ----------------Sheet for AOUT component----------------
        for i in range(len(df)):
            # if str(df.loc[i, "Sort"]) == "1" and pd.notna(df.loc[i, "Component Description"]): area = df.loc[i, "Component Description"]
            identifier = df.loc[i,"CM Type"]
            if identifier == ("AOUT"):
                tagPart2 = df.loc[i, "TAG Part 2 (# Item)"]
                tagPart2_noDot = tagPart2.split('.', 1)[0]
                aout_dict = {'Tag Part 1'                : df.loc[i, "TAG Part 1 (# System)"],
                            'Tag Part 2'                : df.loc[i, "TAG Part 2 (# Item)"],
                            'Tag Part 3'                : df.loc[i, "TAG Part 3 (Function code)"],
                            'Tag Part 4'                : df.loc[i, "TAG Part 4 (# Function serial)"],
                            'Tag'                       : f'{tagPart2_noDot}_{df.loc[i, "TAG Part 3 (Function code)"]}_{df.loc[i, "TAG Part 4 (# Function serial)"]}',
                            'Description'               : f'{df.loc[i, "Component Description"]} - {df.loc[i, "Process Description"]}',
                            'CM Type'                   : df.loc[i,"CM Type"],
                            'CM No'                     : int(df.loc[i, "CM No"]),
                            'Cabinet'                   : df.loc[i, "TAG Part 2 (# Item)"],
                            'Area'                      : f'{df.loc[i, "Tab ID"]}{df.loc[i, "Sub ID"]}',
                            'PLC No.'                   : df.loc[i, "PLC NO"],
                            'Tab ID'                    : df.loc[i, "Tab ID"],
                            'Sub ID'                    : df.loc[i, "Sub ID"],
                            'SC_Local'                  : df.loc[i, "SC Local"],
                            'PLC Tag Name'              : df.loc[i, "PLC Tag Name"],
                            'PLC Tag Address'           : df.loc[i, "PLC Address"],
                            'EU'                        : df.loc[i, "EU"],
                            'Min EU'                    : df.loc[i, "Min EU"],
                            'Max EU'                    : df.loc[i, "Max EU"],
                            'Raw Min'                   : df.loc[i, "Raw Min"],
                            'Raw Max'                   : df.loc[i, "Raw Max"],
                            'Pdf Elect Schema Page'     : df.loc[i, "Pdf Elect Schema Page"],
                            'Pdf Vfd Schema Page'       : df.loc[i, "Pdf Vfd Schema Page"],
                            'Checked'                   : ''
                            }
                aout_list.append(aout_dict)
        
        return aout_list

    def CreatePIDSheet(pid_list): # ----------------Sheet for PID component----------------
        for i in range(len(df)):
            # if str(df.loc[i, "Sort"]) == "1" and pd.notna(df.loc[i, "Component Description"]): area = df.loc[i, "Component Description"]
            identifier = df.loc[i,"CM Type"]
            if identifier == ("PID"):
                tagPart2 = df.loc[i, "TAG Part 2 (# Item)"]
                tagPart2_noDot = tagPart2.split('.', 1)[0]
                pid_dict = {'Tag Part 1'                : df.loc[i, "TAG Part 1 (# System)"],
                            'Tag Part 2'                : df.loc[i, "TAG Part 2 (# Item)"],
                            'Tag Part 3'                : df.loc[i, "TAG Part 3 (Function code)"],
                            'Tag Part 4'                : df.loc[i, "TAG Part 4 (# Function serial)"],
                            'Tag'                       : f'{tagPart2_noDot}_{df.loc[i, "TAG Part 3 (Function code)"]}_{df.loc[i, "TAG Part 4 (# Function serial)"]}',
                            'Description'               : f'{df.loc[i, "Component Description"]} - {df.loc[i, "Process Description"]}',
                            'CM Type'                   : df.loc[i,"CM Type"],
                            'CM No'                     : int(df.loc[i, "CM No"]),
                            'Cabinet'                   : df.loc[i, "TAG Part 2 (# Item)"],
                            'Area'                      : f'{df.loc[i, "Tab ID"]}{df.loc[i, "Sub ID"]}',
                            'PLC No.'                   : df.loc[i, "PLC NO"],
                            'Tab ID'                    : df.loc[i, "Tab ID"],
                            'Sub ID'                    : df.loc[i, "Sub ID"],
                            'SC_Local'                  : df.loc[i, "SC Local"],
                            'PLC Tag Name'              : df.loc[i, "PLC Tag Name"],
                            'PLC Tag Address'           : df.loc[i, "PLC Address"],
                            'EU'                        : df.loc[i, "EU"],
                            'Min EU'                    : df.loc[i, "Min EU"],
                            'Max EU'                    : df.loc[i, "Max EU"],
                            'Raw Min'                   : df.loc[i, "Raw Min"],
                            'Raw Max'                   : df.loc[i, "Raw Max"],
                            'Pdf Elect Schema Page'     : df.loc[i, "Pdf Elect Schema Page"],
                            'Pdf Vfd Schema Page'       : df.loc[i, "Pdf Vfd Schema Page"],
                            'Checked'                   : ''
                            }
                pid_list.append(pid_dict)
        
        return pid_list

    def CreateALMSheet(alm_list): # ----------------Sheet for ALM component----------------
        for i in range(len(df)):
            # if str(df.loc[i, "Sort"]) == "1" and pd.notna(df.loc[i, "Component Description"]): area = df.loc[i, "Component Description"]
            identifier = df.loc[i,"CM Type"]
            if identifier == ("ALM"):
                tagPart2 = df.loc[i, "TAG Part 2 (# Item)"]
                tagPart2_noDot = tagPart2.split('.', 1)[0]
                alm_dict = {'Tag Part 1'                : df.loc[i, "TAG Part 1 (# System)"],
                            'Tag Part 2'                : df.loc[i, "TAG Part 2 (# Item)"],
                            'Tag Part 3'                : df.loc[i, "TAG Part 3 (Function code)"],
                            'Tag Part 4'                : df.loc[i, "TAG Part 4 (# Function serial)"],
                            'Tag'                       : f'{tagPart2_noDot}_{df.loc[i, "TAG Part 3 (Function code)"]}_{df.loc[i, "TAG Part 4 (# Function serial)"]}',
                            'Description'               : f'{df.loc[i, "Component Description"]} - {df.loc[i, "Process Description"]}',
                            'CM Type'                   : df.loc[i,"CM Type"],
                            'CM No'                     : int(df.loc[i, "CM No"]),
                            'Cabinet'                   : df.loc[i, "TAG Part 2 (# Item)"],
                            'Area'                      : f'{df.loc[i, "Tab ID"]}{df.loc[i, "Sub ID"]}',
                            'PLC No.'                   : df.loc[i, "PLC NO"],
                            'Tab ID'                    : df.loc[i, "Tab ID"],
                            'Sub ID'                    : df.loc[i, "Sub ID"],
                            'SC_Local'                  : df.loc[i, "SC Local"],
                            'Cfg Type'                  : df.loc[i, "Cfg Type"],
                            'Cfg Invert'                : df.loc[i, "Cfg Invert"],
                            'PLC Tag Name'              : df.loc[i, "PLC Tag Name"],
                            'PLC Tag Address'           : df.loc[i, "PLC Address"],
                            'Pdf Elect Schema Page'     : df.loc[i, "Pdf Elect Schema Page"],
                            'Pdf Vfd Schema Page'       : df.loc[i, "Pdf Vfd Schema Page"],
                            'Checked'                   : ''
                            }
                alm_list.append(alm_dict)
        
        return alm_list

    def CreateFPSheet(fp_list): # ----------------Sheet for FP component----------------
        for i in range(len(df)):
            # if str(df.loc[i, "Sort"]) == "1" and pd.notna(df.loc[i, "Component Description"]): area = df.loc[i, "Component Description"]
            identifier = df.loc[i,"CM Type"]
            if identifier in identifiers:
                tagPart2 = df.loc[i, "TAG Part 2 (# Item)"]
                tagPart2_noDot = tagPart2.split('.', 1)[0]
                fp_dict = {'Tag Part 1'                : df.loc[i, "TAG Part 1 (# System)"],
                            'Tag Part 2'                : df.loc[i, "TAG Part 2 (# Item)"],
                            'Tag Part 3'                : df.loc[i, "TAG Part 3 (Function code)"],
                            'Tag Part 4'                : df.loc[i, "TAG Part 4 (# Function serial)"],
                            'Tag'                       : f'{tagPart2_noDot}_{df.loc[i, "TAG Part 3 (Function code)"]}_{df.loc[i, "TAG Part 4 (# Function serial)"]}',
                            'Description'               : f'{df.loc[i, "Component Description"]} - {df.loc[i, "Process Description"]}',
                            'CM Type'                   : df.loc[i,"CM Type"],
                            'CM No'                     : int(df.loc[i, "CM No"]),
                            'Cabinet'                   : df.loc[i, "TAG Part 2 (# Item)"],
                            'Area'                      : f'{df.loc[i, "Tab ID"]}{df.loc[i, "Sub ID"]}',
                            'PLC No.'                   : df.loc[i, "PLC NO"],
                            'Tab ID'                    : df.loc[i, "Tab ID"],
                            'Sub ID'                    : df.loc[i, "Sub ID"],
                            'SC_Local'                  : df.loc[i, "SC Local"],
                            'PLC Tag Name'              : df.loc[i, "PLC Tag Name"],
                            'PLC Tag Address'           : df.loc[i, "PLC Address"],
                            'EU'                        : df.loc[i, "EU"],
                            }
                fp_list.append(fp_dict)
        
        return fp_list
    
    def CreateFPVLVSheet(fp_vlv_list): # ----------------Sheet for FP_VLV component----------------
        for i in range(len(df)):
            # if str(df.loc[i, "Sort"]) == "1" and pd.notna(df.loc[i, "Component Description"]): area = df.loc[i, "Component Description"]
            identifier = df.loc[i,"CM Type"]
            if identifier == ("VLV"):
                tagPart2 = df.loc[i, "TAG Part 2 (# Item)"]
                tagPart2_noDot = tagPart2.split('.', 1)[0]
                fp_vlv_dict = {'Tag Part 1'                : df.loc[i, "TAG Part 1 (# System)"],
                                'Tag Part 2'                : df.loc[i, "TAG Part 2 (# Item)"],
                                'Tag Part 3'                : df.loc[i, "TAG Part 3 (Function code)"],
                                'Tag Part 4'                : df.loc[i, "TAG Part 4 (# Function serial)"],
                                'Tag'                       : f'{tagPart2_noDot}_{df.loc[i, "TAG Part 3 (Function code)"]}_{df.loc[i, "TAG Part 4 (# Function serial)"]}',
                                'Description'               : f'{df.loc[i, "Component Description"]} - {df.loc[i, "Process Description"]}',
                                'CM Type'                   : df.loc[i,"CM Type"],
                                'CM No'                     : int(df.loc[i, "CM No"]),
                                'Cabinet'                   : df.loc[i, "TAG Part 2 (# Item)"],
                                'Area'                      : f'{df.loc[i, "Tab ID"]}{df.loc[i, "Sub ID"]}',
                                'PLC No.'                   : df.loc[i, "PLC NO"],
                                'Tab ID'                    : df.loc[i, "Tab ID"],
                                'Sub ID'                    : df.loc[i, "Sub ID"],
                                'SC_Local'                  : df.loc[i, "SC Local"],
                                'PLC Tag Name'              : df.loc[i, "PLC Tag Name"],
                                'PLC Tag Address'           : df.loc[i, "PLC Address"],
                                'EU'                        : df.loc[i, "EU"],
                                'Valve Out P'               : '',
                                'Valve Out N'               : '',
                                'Valve FB P'                : '',
                                'Valve FB N'                : '',
                                'Activate Text'             : df.loc[i, "Activate Text"],
                                'Deactivate Text'           : df.loc[i, "Deactivate Text"]
                            }
                fp_vlv_list.append(fp_vlv_dict)
        
        return fp_vlv_list

    def CreateSAFESheet(aoi_safe_list): # ----------------Sheet for SAFE component----------------
        for i in range(len(df)):
            # if str(df.loc[i, "Sort"]) == "1" and pd.notna(df.loc[i, "Component Description"]): area = df.loc[i, "Component Description"]
            identifier = df.loc[i,"CM Type"]
            signal_type1 = df.loc[i, "Signal Type"] == "SDI"
            signal_type2 = df.loc[i, "Signal Type"] == "SDO"
            sort_nr1 = df.loc[i, "Sort"] == 3
            sort_nr2 = df.loc[i, "Sort"] == "3"
            if identifier in identifiers and (signal_type1 or signal_type2) and (sort_nr1 or sort_nr2):
                tagPart2 = df.loc[i, "TAG Part 2 (# Item)"]
                tagPart2_noDot = tagPart2.split('.', 1)[0]
                aoi_safe_dict = {'Tag Part 1'                : df.loc[i, "TAG Part 1 (# System)"],
                                'Tag Part 2'                : df.loc[i, "TAG Part 2 (# Item)"],
                                'Tag Part 3'                : df.loc[i, "TAG Part 3 (Function code)"],
                                'Tag Part 4'                : df.loc[i, "TAG Part 4 (# Function serial)"],
                                'Tag'                       : f'{tagPart2_noDot}_{df.loc[i, "TAG Part 3 (Function code)"]}_{df.loc[i, "TAG Part 4 (# Function serial)"]}',
                                'Description'               : f'{df.loc[i, "Component Description"]} - {df.loc[i, "Process Description"]}',
                                'CM Type'                   : df.loc[i,"CM Type"],
                                'CM No'                     : int(df.loc[i, "CM No"]),
                                'Cabinet'                   : df.loc[i, "TAG Part 2 (# Item)"],
                                'Area'                      : f'{df.loc[i, "Tab ID"]}{df.loc[i, "Sub ID"]}',
                                'PLC No.'                   : df.loc[i, "PLC NO"],
                                'Tab ID'                    : df.loc[i, "Tab ID"],
                                'Sub ID'                    : df.loc[i, "Sub ID"],
                                'SC_Local'                  : df.loc[i, "SC Local"],
                                'PLC Tag Name'              : df.loc[i, "PLC Tag Name"],
                                'PLC Tag Address'           : df.loc[i, "PLC Address"]
                            }
                aoi_safe_list.append(aoi_safe_dict)

        return aoi_safe_list

    def CreateSUMSheet(sum_list): # ----------------Sheet for SUM - Power Monitor component----------------
        for i in range(len(df)):
            # if str(df.loc[i, "Sort"]) == "1" and pd.notna(df.loc[i, "Component Description"]): area = df.loc[i, "Component Description"]
            identifier = df.loc[i,"CM Type"]
            if identifier == ("SUM"):
                tagPart2 = df.loc[i, "TAG Part 2 (# Item)"]
                tagPart2_noDot = tagPart2.split('.', 1)[0]
                sum_dict = {'Tag Part 1'                : df.loc[i, "TAG Part 1 (# System)"],
                                'Tag Part 2'                : df.loc[i, "TAG Part 2 (# Item)"],
                                'Tag Part 3'                : df.loc[i, "TAG Part 3 (Function code)"],
                                'Tag Part 4'                : df.loc[i, "TAG Part 4 (# Function serial)"],
                                'Tag'                       : f'{tagPart2_noDot}_{df.loc[i, "TAG Part 3 (Function code)"]}_{df.loc[i, "TAG Part 4 (# Function serial)"]}',
                                'Description'               : f'{df.loc[i, "Component Description"]} - {df.loc[i, "Process Description"]}',
                                'CM Type'                   : df.loc[i,"CM Type"],
                                'CM No'                     : int(df.loc[i, "CM No"]),
                                'Cabinet'                   : df.loc[i, "TAG Part 2 (# Item)"],
                                'Area'                      : f'{df.loc[i, "Tab ID"]}{df.loc[i, "Sub ID"]}',
                                'PLC No.'                   : df.loc[i, "PLC NO"],
                                'Tab ID'                    : df.loc[i, "Tab ID"],
                                'Sub ID'                    : df.loc[i, "Sub ID"],
                                'SC_Local'                  : df.loc[i, "SC Local"],
                                'IP Address'                : df.loc[i, "Net number / IP address"],
                            }
                sum_list.append(sum_dict)
        
        return sum_list

    def CreateIPSheet(ipAdr_list): # ----------------Sheet for IP Addresses ----------------
        for i in range(len(df)):
            # if str(df.loc[i, "Sort"]) == "1" and pd.notna(df.loc[i, "Component Description"]): area = df.loc[i, "Component Description"]
            if pd.notna(df.loc[i, "Net number / IP address"]):
                tagPart2 = df.loc[i, "TAG Part 2 (# Item)"]
                tagPart2_noDot = tagPart2.split('.', 1)[0]
                tagPart3 = df.loc[i, "TAG Part 3 (Function code)"]
                #tagPart3 = tagPart3.replace("TA", "MA")
                identifier = df.loc[i,"CM Type"]
                if identifier == "FRQ_ETH": identifier = "FRQ"
                ipAdr_dict = {'Tag Part 1'                : df.loc[i, "TAG Part 1 (# System)"],
                                'Tag Part 2'                : df.loc[i, "TAG Part 2 (# Item)"],
                                'Tag Part 3'                : df.loc[i, "TAG Part 3 (Function code)"],
                                'Tag Part 4'                : df.loc[i, "TAG Part 4 (# Function serial)"],
                                'Tag'                       : f'{tagPart2_noDot}_{df.loc[i, "TAG Part 3 (Function code)"]}_{df.loc[i, "TAG Part 4 (# Function serial)"]}',
                                'Description'               : f'{df.loc[i, "Component Description"]} - {df.loc[i, "Process Description"]}',
                                'CM Type'                   : df.loc[i,"CM Type"],
                                'CM No'                     : int(df.loc[i, "CM No"]),
                                'Cabinet'                   : df.loc[i, "TAG Part 2 (# Item)"],
                                'Area'                      : f'{df.loc[i, "Tab ID"]}{df.loc[i, "Sub ID"]}',
                                'PLC No.'                   : df.loc[i, "PLC NO"],
                                'IP-address'                : df.loc[i, "Net number / IP address"],
                                'Tag Name'                  : f'{tagPart2_noDot}_{df.iloc[i, 3]}_{tagPart3}_{identifier}_UDT'
                            }
                ipAdr_list.append(ipAdr_dict)

        return ipAdr_list

    def CreateHWSheet(hw_list): # ----------------Sheet for HW Configuration ----------------
        for i in range(len(df)):
            hwIDs = ["DI", "DO", "ETH", "SDO", "SDI", "AI", "AO"]
            identifier = df.loc[i,"Signal Type"]
            if identifier in hwIDs:
                hw_dict = {'Device Name'                : df.loc[i, "Component Description"],
                            'Device Item Name'          : df.loc[i, "TAG Part 4 (# Function serial)"],
                            'Type'                      : df.loc[i, "Process Description"],
                            'Slot'                      : df.loc[i, "PLC Address"],
                            'Order Number'              : df.loc[i, "Component remark"],
                            'Version'                   : df.loc[i, "FW Version"],
                            'IP-address'                : df.loc[i, "Net number / IP address"],
                            'IO-address'                : df.loc[i, "PLC Bit address"]
                            }
                hw_list.append(hw_dict)

        return hw_list
    
    def CreateAreaSheet(area_list): # ----------------Sheet for Area ----------------
        def isNaN(string):
            return string != string

        areaNames = []
        finalArea = str()

        for i in range(len(df)):
            if str(df.loc[i, "Sort"]) == "1" and pd.notna(df.loc[i, "Component Description"]):
                area1 = df.loc[i, "Component Description"]
                area2 = df.loc[i, "Component Description"]
                area1Sort = df.loc[i, "Sort"]
                area2_sort = df.loc[i+1, "Sort"]
                if isNaN(area2):
                    finalArea = area1
                elif not isNaN(area2) and area2_sort != 1:
                    finalArea = area1
                elif not isNaN(area2) and area2_sort == 1:
                    finalArea = area2
            if finalArea not in areaNames and finalArea != "":
                areaNames.append(finalArea)
                area_dict = {'Area ID'    : i,
                            'Area Name'   : finalArea
                            }
                area_list.append(area_dict)

        return area_list
    
    def CreateCableSheet(cable_list): # ----------------Sheet for HW Configuration ----------------
        for i in range(len(df)):
            if pd.notna(df.loc[i, "Cable Plan"]):
                tagPart2 = df.loc[i, "TAG Part 2 (# Item)"]
                tagPart2_noDot = tagPart2.split('.', 1)[0]
                cable_dict = {'Tag Part 1'              : df.loc[i, "TAG Part 1 (# System)"],
                            'Tag Part 2'                : df.loc[i, "TAG Part 2 (# Item)"],
                            'Tag Part 3'                : df.loc[i, "TAG Part 3 (Function code)"],
                            'Tag Part 4'                : df.loc[i, "TAG Part 4 (# Function serial)"],
                            'Tag'                       : f'{tagPart2_noDot}_{df.loc[i, "TAG Part 3 (Function code)"]}_{df.loc[i, "TAG Part 4 (# Function serial)"]}',
                            'Description'               : f'{df.loc[i, "Component Description"]} - {df.loc[i, "Process Description"]}',
                            'CM Type'                   : df.loc[i,"CM Type"],
                            'CM No'                     : int(df.loc[i, "CM No"]),
                            'Cabinet'                   : df.loc[i, "TAG Part 2 (# Item)"],
                            'Area'                      : f'{df.loc[i, "Tab ID"]}{df.loc[i, "Sub ID"]}',
                            'Cable Plan'                : df.loc[i, "Cable Plan"],
                            'Cable Type'                : df.loc[i, "Cable Type"],
                            'Cable'                     : df.loc[i, "Cable"],
                            'Dimensions'                : df.loc[i, "Number of cores"],
                            'Cable Name'                : f'+{df.loc[i, "TAG Part 2 (# Item)"]}={df.loc[i, "TAG Part 3 (Function code)"]}-{df.loc[i, "Cable Type"]}{df.loc[i, "TAG Part 4 (# Function serial)"][2:]}'
                            }
                cable_list.append(cable_dict)

        return cable_list


    #################################### Run Functions depending on which User Settings are Selected ####################################
    if Cmd_AssignCMNo:
        try: AssignCMNumber() 
        except: Sts_PrgStatus = "Assigning Control Module number failed"
    if Cmd_CreateSheets:
        print("Create Sheets") # Test code
        try: CreateFRQSheet(frq_list)
        except: Sts_PrgStatus = "Creating FRQ Sheet Failed"
        try: CreateMTRSheet(mtr_list)
        except: Sts_PrgStatus = "Creating MTR Sheet Failed"
        try: CreateVLVSheet(vlv_list)
        except: Sts_PrgStatus = "Creating VLV Sheet Failed"
        try: CreateDINSheet(din_list)
        except: Sts_PrgStatus = "Creating DIN Sheet Failed"
        try: CreateAINSheet(ain_list)
        except: Sts_PrgStatus = "Creating AIN Sheet Failed"
        try: CreateAOUTSheet(aout_list)
        except: Sts_PrgStatus = "Creating AOUT Sheet Failed"
        try: CreatePIDSheet(pid_list)
        except: Sts_PrgStatus = "Creating PID Sheet Failed"
        try: CreateALMSheet(alm_list)
        except: Sts_PrgStatus = "Creating ALM Sheet Failed"
        try: CreateFPSheet(fp_list)
        except: Sts_PrgStatus = "Creating FP Sheet Failed"
        try: CreateFPVLVSheet(fp_vlv_list)
        except: Sts_PrgStatus = "Creating FP_VLV Sheet Failed"
        try: CreateSAFESheet(aoi_safe_list)
        except: Sts_PrgStatus = "Creating SAFE Sheet Failed"
        try: CreateSUMSheet(sum_list)
        except: Sts_PrgStatus = "Creating SUM Sheet Failed"
        try: CreateIPSheet(ipAdr_list)
        except: Sts_PrgStatus = "Creating IP Sheet Failed"
        try: CreateHWSheet(hw_list)
        except: Sts_PrgStatus = "Creating HW Sheet Failed"
        try: CreateAreaSheet(area_list)
        except: Sts_PrgStatus = "Creating Area Sheet Failed"
        try: CreateCableSheet(cable_list)
        except: Sts_PrgStatus = "Creating Cable Sheet Failed"


    # Create dataframes to be exported to sheets
    frq_df = pd.DataFrame(frq_list)
    mtr_df = pd.DataFrame(mtr_list)
    vlv_df = pd.DataFrame(vlv_list)
    din_df = pd.DataFrame(din_list)
    ain_df = pd.DataFrame(ain_list)
    aout_df = pd.DataFrame(aout_list)
    pid_df = pd.DataFrame(pid_list)
    alm_df = pd.DataFrame(alm_list)
    fp_df = pd.DataFrame(fp_list)
    fp_vlv_df = pd.DataFrame(fp_vlv_list)
    aoi_safe_df = pd.DataFrame(aoi_safe_list)
    sum_df = pd.DataFrame(sum_list)
    ipAdr_df = pd.DataFrame(ipAdr_list)
    hw_df = pd.DataFrame(hw_list)
    area_df = pd.DataFrame(area_list)
    cable_df = pd.DataFrame(cable_list)

    # Sort rows in dataframes by CM No
    if Cmd_SortRows:
        try:
            if not frq_df.empty: frq_df = frq_df.sort_values(by ='CM No')
            if not mtr_df.empty: mtr_df = mtr_df.sort_values(by ='CM No')
            if not vlv_df.empty: vlv_df = vlv_df.sort_values(by ='CM No')
            if not din_df.empty: din_df = din_df.sort_values(by ='CM No')
            if not ain_df.empty: ain_df = ain_df.sort_values(by ='CM No')
            if not aout_df.empty: aout_df = aout_df.sort_values(by ='CM No')
            if not pid_df.empty: pid_df = pid_df.sort_values(by ='CM No')
            if not alm_df.empty: alm_df = alm_df.sort_values(by ='CM No')
            if not fp_df.empty: fp_df = fp_df.sort_values(by ='CM No')
            if not fp_vlv_df.empty: fp_vlv_df = fp_vlv_df.sort_values(by ='CM No')
            if not aoi_safe_df.empty: aoi_safe_df = aoi_safe_df.sort_values(by ='CM No')
            if not sum_df.empty: sum_df = sum_df.sort_values(by ='CM No')
            if not ipAdr_df.empty: ipAdr_df = ipAdr_df.sort_values(by ='CM No')
            if not cable_df.empty: cable_df = cable_df.sort_values(by ='CM No')
        except:
            Sts_PrgStatus = "Sorting rows failed"

    # Check for duplicated CM No
    if Cmd_CheckDups:
        try:
            if not frq_df.empty: frq_df['Duplicates'] = np.where(frq_df['CM No'] == frq_df['CM No'].shift(), 'Duplicated', 'Not Duplicated')
            if not mtr_df.empty: mtr_df['Duplicates'] = np.where(mtr_df['CM No'] == mtr_df['CM No'].shift(), 'Duplicated', 'Not Duplicated')
            if not vlv_df.empty: vlv_df['Duplicates'] = np.where(vlv_df['CM No'] == vlv_df['CM No'].shift(), 'Duplicated', 'Not Duplicated')
            if not din_df.empty: din_df['Duplicates'] = np.where(din_df['CM No'] == din_df['CM No'].shift(), 'Duplicated', 'Not Duplicated')
            if not ain_df.empty: ain_df['Duplicates'] = np.where(ain_df['CM No'] == ain_df['CM No'].shift(), 'Duplicated', 'Not Duplicated')
            if not aout_df.empty: aout_df['Duplicates'] = np.where(aout_df['CM No'] == aout_df['CM No'].shift(), 'Duplicated', 'Not Duplicated')
            if not pid_df.empty: pid_df['Duplicates'] = np.where(pid_df['CM No'] == pid_df['CM No'].shift(), 'Duplicated', 'Not Duplicated')
            if not alm_df.empty: alm_df['Duplicates'] = np.where(alm_df['CM No'] == alm_df['CM No'].shift(), 'Duplicated', 'Not Duplicated')
            if not aoi_safe_df.empty: aoi_safe_df['Duplicates'] = np.where(aoi_safe_df['CM No'] == aoi_safe_df['CM No'].shift(), 'Duplicated', 'Not Duplicated')
            if not sum_df.empty: sum_df['Duplicates'] = np.where(sum_df['CM No'] == sum_df['CM No'].shift(), 'Duplicated', 'Not Duplicated')
            if not ipAdr_df.empty: ipAdr_df['Duplicates'] = np.where(ipAdr_df['CM No'] == ipAdr_df['CM No'].shift(), 'Duplicated', 'Not Duplicated')
        except:
            Sts_PrgStatus = "Checking duplicated control module numbers failed"


    # ----------------Create all sheets and make column width dynamic----------------
    filename = "export/worksheet/Siemens_Worksheet.xlsx"

    def get_col_widths(dataframe):
        return [max([len(str(s)) for s in dataframe[col].values] + [len(col)]) for col in dataframe.columns]

    # If a cell is empty, Pandas interprets the cell as a float. To avoid this, fill empty cells.
    df.replace(np.nan, "", inplace=True)
    frq_df.replace(np.nan, "", inplace=True)
    mtr_df.replace(np.nan, "", inplace=True)
    vlv_df.replace(np.nan, "", inplace=True)
    din_df.replace(np.nan, "", inplace=True)
    ain_df.replace(np.nan, "", inplace=True)
    aout_df.replace(np.nan, "", inplace=True)
    pid_df.replace(np.nan, "", inplace=True)
    alm_df.replace(np.nan, "", inplace=True)
    fp_df.replace(np.nan, "", inplace=True)
    fp_vlv_df.replace(np.nan, "", inplace=True)
    aoi_safe_df.replace(np.nan, "", inplace=True)
    sum_df.replace(np.nan, "", inplace=True)
    ipAdr_df.replace(np.nan, "", inplace=True)
    hw_df.replace(np.nan, "", inplace=True)
    area_df.replace(np.nan, "", inplace=True)
    cable_df.replace(np.nan, "", inplace=True)

    # # Create sheets
    if Cmd_CreateSheets:
        with pd.ExcelWriter(filename, engine='xlsxwriter') as writer:
            df.to_excel(writer, sheet_name='IO List', index=False)
            frq_df.to_excel(writer, sheet_name='FRQ', index=False)
            mtr_df.to_excel(writer, sheet_name='MTR', index=False)
            vlv_df.to_excel(writer, sheet_name='VLV', index=False)
            din_df.to_excel(writer, sheet_name='DIN', index=False)
            ain_df.to_excel(writer, sheet_name='AIN', index=False)
            aout_df.to_excel(writer, sheet_name='AOUT', index=False)
            pid_df.to_excel(writer, sheet_name='PID', index=False)
            alm_df.to_excel(writer, sheet_name='ALM', index=False)
            fp_df.to_excel(writer, sheet_name='FP', index=False)
            fp_vlv_df.to_excel(writer, sheet_name='FP_VLV', index=False)
            aoi_safe_df.to_excel(writer, sheet_name='SAFE', index=False)
            sum_df.to_excel(writer, sheet_name='SUM', index=False)
            ipAdr_df.to_excel(writer, sheet_name='IP Address', index=False)
            hw_df.to_excel(writer, sheet_name='HW', index=False)
            area_df.to_excel(writer, sheet_name='Area', index=False)
            cable_df.to_excel(writer, sheet_name='Cable Plan', index=False)
            for i, width in enumerate(get_col_widths(df)):
                    writer.sheets['IO List'].set_column(i, i, width)    
            for i, width in enumerate(get_col_widths(frq_df)):
                    writer.sheets['FRQ'].set_column(i, i, width)
            for i, width in enumerate(get_col_widths(mtr_df)):
                    writer.sheets['MTR'].set_column(i, i, width)
            for i, width in enumerate(get_col_widths(vlv_df)):
                    writer.sheets['VLV'].set_column(i, i, width)
            for i, width in enumerate(get_col_widths(din_df)):
                    writer.sheets['DIN'].set_column(i, i, width)
            for i, width in enumerate(get_col_widths(ain_df)):
                    writer.sheets['AIN'].set_column(i, i, width)
            for i, width in enumerate(get_col_widths(aout_df)):
                    writer.sheets['AOUT'].set_column(i, i, width)
            for i, width in enumerate(get_col_widths(pid_df)):
                    writer.sheets['PID'].set_column(i, i, width)
            for i, width in enumerate(get_col_widths(alm_df)):
                    writer.sheets['ALM'].set_column(i, i, width)
            for i, width in enumerate(get_col_widths(fp_df)):
                    writer.sheets['FP'].set_column(i, i, width)
            for i, width in enumerate(get_col_widths(fp_vlv_df)):
                    writer.sheets['FP_VLV'].set_column(i, i, width)
            for i, width in enumerate(get_col_widths(aoi_safe_df)):
                    writer.sheets['SAFE'].set_column(i, i, width)
            for i, width in enumerate(get_col_widths(sum_df)):
                    writer.sheets['SUM'].set_column(i, i, width)
            for i, width in enumerate(get_col_widths(ipAdr_df)):
                    writer.sheets['IP Address'].set_column(i, i, width)
            for i, width in enumerate(get_col_widths(hw_df)):
                    writer.sheets['HW'].set_column(i, i, width)
            for i, width in enumerate(get_col_widths(area_df)):
                    writer.sheets['Area'].set_column(i, i, width)
            for i, width in enumerate(get_col_widths(cable_df)):
                    writer.sheets['Cable Plan'].set_column(i, i, width)
    
    Sts_PrgStatus = "Completed"

def StyleWS():
    filepath = r'export/worksheet/Siemens_Worksheet.xlsx'
    wb = load_workbook(filepath)

    bg_color = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    border_color = Border(right=Side(style='thin'))
    font_color = Font(color='000000', bold=True)


# ----------------Apply styles to sheets----------------
    ws = wb['IO List']
    row_count = ws.max_row + 1
    for x in range(row_count):
        if x % 2 != 0:
            for cell in ws[x:x]:
                cell.fill = bg_color
    
    ws = wb['FRQ']
    row_count = ws.max_row + 1
    for x in range(row_count):
        if x % 2 != 0:
            for cell in ws[x:x]:
                cell.fill = bg_color
    
    ws = wb['MTR']
    row_count = ws.max_row + 1
    for x in range(row_count):
        if x % 2 != 0:
            for cell in ws[x:x]:
                cell.fill = bg_color
    
    ws = wb['VLV']
    row_count = ws.max_row + 1
    for x in range(row_count):
        if x % 2 != 0:
            for cell in ws[x:x]:
                cell.fill = bg_color
    
    ws = wb['DIN']
    row_count = ws.max_row + 1
    for x in range(row_count):
        if x % 2 != 0:
            for cell in ws[x:x]:
                cell.fill = bg_color
    
    ws = wb['AIN']
    row_count = ws.max_row + 1
    for x in range(row_count):
        if x % 2 != 0:
            for cell in ws[x:x]:
                cell.fill = bg_color
    
    ws = wb['AOUT']
    row_count = ws.max_row + 1
    for x in range(row_count):
        if x % 2 != 0:
            for cell in ws[x:x]:
                cell.fill = bg_color
    
    ws = wb['PID']
    row_count = ws.max_row + 1
    for x in range(row_count):
        if x % 2 != 0:
            for cell in ws[x:x]:
                cell.fill = bg_color
    
    ws = wb['ALM']
    row_count = ws.max_row + 1
    for x in range(row_count):
        if x % 2 != 0:
            for cell in ws[x:x]:
                cell.fill = bg_color
    
    ws = wb['FP']
    row_count = ws.max_row + 1
    for x in range(row_count):
        if x % 2 != 0:
            for cell in ws[x:x]:
                cell.fill = bg_color
    
    ws = wb['FP_VLV']
    row_count = ws.max_row + 1
    for x in range(row_count):
        if x % 2 != 0:
            for cell in ws[x:x]:
                cell.fill = bg_color
    
    ws = wb['SAFE']
    row_count = ws.max_row + 1
    for x in range(row_count):
        if x % 2 != 0:
            for cell in ws[x:x]:
                cell.fill = bg_color
    
    ws = wb['SUM']
    row_count = ws.max_row + 1
    for x in range(row_count):
        if x % 2 != 0:
            for cell in ws[x:x]:
                cell.fill = bg_color
    
    ws = wb['IP Address']
    row_count = ws.max_row + 1
    for x in range(row_count):
        if x % 2 != 0:
            for cell in ws[x:x]:
                cell.fill = bg_color
    
    ws = wb['HW']
    row_count = ws.max_row + 1
    for x in range(row_count):
        if x % 2 != 0:
            for cell in ws[x:x]:
                cell.fill = bg_color
    
    ws = wb['Area']
    row_count = ws.max_row + 1
    for x in range(row_count):
        if x % 2 != 0:
            for cell in ws[x:x]:
                cell.fill = bg_color
    
    ws = wb['Cable Plan']
    row_count = ws.max_row + 1
    for x in range(row_count):
        if x % 2 != 0:
            for cell in ws[x:x]:
                cell.fill = bg_color

    

    wb.save(filepath)
    wb.close()