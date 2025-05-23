from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.utils import get_column_letter
import pandas as pd

# import updateAreaNames

def updateIOList(filepath, Cmd_ColorRowsFromSort, Cmd_MarkDupAddressIO, Cmd_MarkDupAddressIP, Cmd_CreateIOTags_Sie, Cmd_CreateIOTags_RA, Cmd_UpdateMachineAreas):

    wb = load_workbook(filepath)
    ws = wb['IO-list']

    row_count = ws.max_row
    column_count = ws.max_column

    Sts_PrgStatus = "Begin"
    
    # ----------------Defining constants----------------
    ioTypes = ['DI', 'DO', 'AI', 'AO', 'SDI', 'SDO']
    digitalIn = 'DI'
    digitalOut = 'DO'
    analogIn = 'AI'
    analogOut = 'AO'
    safeIn = 'SDI'
    safeOut = 'SDO'
    
    # ----------------Creating a header list (below variables represents a column in the IO-list)----------------
    column_list             = [cell.value for cell in ws[4]]
    signalType              = column_list.index('Signal Type')
    plcTagName              = column_list.index('PLC Tag Name')
    tagPart2                = column_list.index('TAG Part 2 (# Item)')
    tagPart3                = column_list.index('TAG Part 3 (Function code)')
    tagPart4                = column_list.index('TAG Part 4 (# Function serial)')
    componentDescription    = column_list.index('Component Description')
    processDescription      = column_list.index('Process Description')
    plcFunction             = column_list.index('PLC Function')
    if Cmd_CreateIOTags_Sie:
        plcAddress          = column_list.index('PLC Address')
        plcBitAddress       = column_list.index('PLC Bit address')
    ipAddress               = column_list.index('Net number / IP address')
    ioAddress               = column_list.index('IO Address')
    sort                    = column_list.index('Sort')
    if Cmd_CreateIOTags_RA:
        slice_number        = column_list.index('Slice number')
        point               = column_list.index('Point')

    # ----------------Styles for spreadsheet----------------
    font = Font(name='Calibri Light',
                size=8,
                bold=False,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='FF000000')
    
    fill = PatternFill(fill_type=None)

    fatfont = Font(name='Calibri Light',
                size=8,
                bold=True,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='FF000000')

    dupfill = PatternFill(fill_type='solid',
                    start_color='00FF0000',
                    end_color='00FF0000')
    
    nrfill1 = PatternFill(fill_type='solid',
                    start_color='00FF8080',
                    end_color='00FF8080')

    nrfill2 = PatternFill(fill_type='solid',
                    start_color='00FFFFCC',
                    end_color='00FFFFCC')

    nrfill3 = PatternFill(fill_type=None)

    nrfill4 = PatternFill(fill_type='solid',
                    start_color='00CCCCFF',
                    end_color='00CCCCFF')

    def CreateIOTags_Sie(): # ----------------Create Siemens IO Tags----------------
        for idx, i in enumerate(ws[get_column_letter(signalType+1)]):
            sort_val = str(ws.cell(row=idx+1, column=sort+1).value)              
            if i.value in ioTypes and sort_val =='3':
                # --- Create PLC Tag Name ---
                signal_type = i.value
                if i.value == safeIn: signal_type = digitalIn
                if i.value == safeOut: signal_type = digitalOut
                tag_part2 = ws.cell(row=idx+1, column=tagPart2+1).value
                tag_Part2_noDot = tag_part2.split('.', 1)[0]
                tag_part3 = ws.cell(row=idx+1, column=tagPart3+1).value
                tag_part4 = ws.cell(row=idx+1, column=tagPart4+1).value
                plc_Function = ws.cell(row=idx+1, column=plcFunction+1).value
                concatTagPlc = f'{tag_Part2_noDot}_{tag_part3}_{tag_part4}_{signal_type}_{plc_Function}'
                activeCellPlc = ws.cell(row=idx+1, column=plcTagName+1, value=concatTagPlc)

                # --- Create IO Address for Siemens ---
                io_type = i.value
                if (i.value == digitalIn) or (i.value == safeIn): io_type = 'I'
                if (i.value == digitalOut) or (i.value == safeOut): io_type = 'Q'
                if (i.value == analogIn): io_type = 'IW'
                if (i.value == analogOut): io_type = 'QW'
                plc_address = ws.cell(row=idx+1, column=plcAddress+1).value
                plc_bit_address = ws.cell(row=idx+1, column=plcBitAddress+1).value
                if io_type == 'IW' or io_type == 'QW':
                    concatTagIO = f'%{io_type}{plc_address}'
                else:
                    concatTagIO = f'%{io_type}{plc_address}.{plc_bit_address}'
                activeCellIO = ws.cell(row=idx+1, column=ioAddress+1, value=concatTagIO)

                # --- Add styles ---
                activeCellPlc.font = font
                activeCellPlc.fill = fill
                activeCellIO.font = font
                activeCellIO.fill = fill

    def CreateIOTags_RA(): # ----------------Create Rockwell IO Tags----------------
        for idx, i in enumerate(ws[get_column_letter(signalType+1)]):
            sort_val = str(ws.cell(row=idx+1, column=sort+1).value)
                # --- Create PLC Tag Name for sort 3---
            if i.value in ioTypes and sort_val =='3':
                signal_type = i.value
                if i.value == safeIn: signal_type = digitalIn
                if i.value == safeOut: signal_type = digitalOut
                tag_part2 = ws.cell(row=idx+1, column=tagPart2+1).value
                tag_Part2_noDot = tag_part2.split('.', 1)[0]
                tag_part3 = ws.cell(row=idx+1, column=tagPart3+1).value
                tag_part4 = ws.cell(row=idx+1, column=tagPart4+1).value
                plc_Function = ws.cell(row=idx+1, column=plcFunction+1).value
                concatTagPlc = f'{tag_Part2_noDot}_{tag_part3}_{tag_part4}_{signal_type}_{plc_Function}'
                activeCellPlc = ws.cell(row=idx+1, column=plcTagName+1, value=concatTagPlc)
                # --- Create IO Address for Rockwell ---
                io_type = i.value
                tag_part2 = ws.cell(row=idx+1, column=tagPart2+1).value
                tag_Part2_noDot = tag_part2.split('.', 1)[0]
                slice_nr = ws.cell(row=idx+1, column=slice_number+1).value
                point_nr = ws.cell(row=idx+1, column=point+1).value
                if (i.value == digitalIn): # Digital input
                    io_type = 'I'
                    concatTagIO = f'{tag_Part2_noDot}_IO:{slice_nr}:{io_type}.{point_nr}'

                if (i.value == safeIn): #Safety input
                    io_type = 'I'
                    concatTagIO = f'{tag_Part2_noDot}_IO:{slice_nr}:{io_type}.Pt0{point_nr}Data'

                if (i.value == digitalOut): #Digital output
                    io_type = 'O'
                    concatTagIO = f'{tag_Part2_noDot}_IO:{slice_nr}:{io_type}.{point_nr}'

                if (i.value == safeOut): #Safety output
                    io_type = 'O'
                    concatTagIO = f'{tag_Part2_noDot}_IO:{slice_nr}:{io_type}.Pt0{point_nr}Data'

                if (i.value == analogIn): #Analog input
                    io_type = 'I' 
                    concatTagIO = f'{tag_Part2_noDot}_IO:{slice_nr}:{io_type}.Ch{point_nr}Data'

                if (i.value == analogOut): #Analog output
                    io_type = 'O'
                    concatTagIO = f'{tag_Part2_noDot}_IO:{slice_nr}:{io_type}.Ch{point_nr}Data'

                activeCellIO = ws.cell(row=idx+1, column=plcAddress+1, value=concatTagIO)

                # --- Add styles ---
                activeCellPlc.font = font
                activeCellPlc.fill = fill
                activeCellIO.font = font
                activeCellIO.fill = fill  

                # --- Create PLC Tag Name for sort 4---
            if sort_val =='4':
                signal_type = i.value
                tag_part2 = ws.cell(row=idx+1, column=tagPart2+1).value
                tag_Part2_noDot = tag_part2.split('.', 1)[0]
                tag_part3 = ws.cell(row=idx+1, column=tagPart3+1).value
                tag_part4 = ws.cell(row=idx+1, column=tagPart4+1).value
                plc_Function = ws.cell(row=idx+1, column=plcFunction+1).value
                concatTagPlc = f'{tag_Part2_noDot}_{tag_part3}_{tag_part4}_{signal_type}_{plc_Function}'
                activeCellPlc = ws.cell(row=idx+1, column=plcTagName+1, value=concatTagPlc)

                # --- Add styles ---
                activeCellPlc.font = font
                activeCellPlc.fill = nrfill4

    
    def MarkDupAddressIO (): # ----------------Mark duplicated IO-addresses----------------
        # --- Generate list of all used addresses ---
        usedAdr_list = []
        for idx, i in enumerate(ws[get_column_letter(ipAddress+1)]):
            actualAdr = i.value
            if actualAdr:
                usedAdr_list.append(actualAdr)

        # --- Generate a list of all duplicated addresses ---
        seen = set()
        dupes = [x for x in usedAdr_list if x in seen or seen.add(x)]

        # --- Add styles to all duplicated addresses
        for idx, i in enumerate(ws[get_column_letter(ipAddress+1)]):
            if i.value in dupes:
                activeCellIO = ws.cell(row=idx+1, column=ioAddress+1)
                activeCellIO.font = fatfont
                activeCellIO.fill = dupfill
    
    def MarkDupAddressIP (): # ----------------Mark duplicated IP-addresses----------------
        # --- Generate list of all used addresses ---
        usedAdr_list = []
        for idx, i in enumerate(ws[get_column_letter(ipAddress+1)]):
            actualAdr = i.value
            if actualAdr:
                usedAdr_list.append(actualAdr)

        # --- Generate a list of all duplicated addresses ---
        seen = set()
        dupes = [x for x in usedAdr_list if x in seen or seen.add(x)]

        # --- Add styles to all duplicated addresses ---
        for idx, i in enumerate(ws[get_column_letter(ipAddress+1)]):
            if i.value in dupes:
                activeCellIO = ws.cell(row=idx+1, column=ipAddress+1)
                activeCellIO.font = fatfont
                activeCellIO.fill = dupfill

    def ColorRowsFromSort(): # ----------------Colorize rows from "Sort" number----------------
        for idx, i in enumerate(ws[get_column_letter(sort+1)]):
            sort_nr = i.value
            if (sort_nr == "1") or (sort_nr == 1): # When Sort is 1
                for rows in ws.iter_rows(min_row=idx+1, max_row=idx+1, max_col=column_count):
                    for cell in rows:
                        cell.fill = nrfill1
                        cell.font = fatfont
            if (sort_nr == "2") or (sort_nr == 2): # When Sort is 2
                for rows in ws.iter_rows(min_row=idx+1, max_row=idx+1, max_col=column_count):
                    for cell in rows:
                        cell.fill = nrfill2
                        cell.font = font
            if (sort_nr == "3") or (sort_nr == 3): # When Sort is 3
                for rows in ws.iter_rows(min_row=idx+1, max_row=idx+1, max_col=column_count):
                    for cell in rows:
                        cell.fill = nrfill3
                        cell.font = font
            if (sort_nr == "4") or (sort_nr == 4): # When Sort is 4
                for rows in ws.iter_rows(min_row=idx+1, max_row=idx+1, max_col=column_count):
                    for cell in rows:
                        cell.fill = nrfill4
                        cell.font = font

    # ----------------Function Calling----------------
    if Cmd_ColorRowsFromSort:
        try: ColorRowsFromSort()
        except: Sts_PrgStatus = "Colorizing rows from Sort failed"

    if Cmd_MarkDupAddressIO:
        try: MarkDupAddressIO()
        except: Sts_PrgStatus = "Marking duplicated IO-addresses failed"

    if Cmd_MarkDupAddressIP:
        try: MarkDupAddressIP()
        except: Sts_PrgStatus = "Marking duplicated IP-addresses failed"

    if Cmd_CreateIOTags_Sie and not Cmd_CreateIOTags_RA:
        try: CreateIOTags_Sie()
        except: Sts_PrgStatus = "Creating Siemens IO tags failed"

    if Cmd_CreateIOTags_RA and not Cmd_CreateIOTags_Sie:
        try: CreateIOTags_RA()
        except: Sts_PrgStatus = "Creating Rockwell IO tags failed"

    wb.save(filepath)
    wb.close()

    if Cmd_UpdateMachineAreas:
        try: updateMachineAreas(filepath)
        except: Sts_PrgStatus = "Updating machine areas failed"

def updateMachineAreas(filepath):
        
    df = pd.read_excel(filepath, 'IO-list')
    df.columns = df.iloc[2].astype(str) # Set the header to be at index 2
    df = pd.DataFrame(df.iloc[3:, 1:]) # Remove the first 3 rows and the first column
    df = df.reset_index()

    def get_col_widths(dataframe):
        return [max([len(str(s)) for s in dataframe[col].values] + [len(col)]) for col in dataframe.columns]

    def isNaN(string):
        return string != string
    areaNames = []
    finalArea = str()
    area_list = list()

    for i in range(len(df)):
        if str(df.loc[i, "Sort"]) == "1" and pd.notna(df.loc[i, "Component Description"]):
            area1 = df.loc[i, "Component Description"]
            area2 = df.loc[i, "Component Description"]
            area1Sort = df.loc[i, "Sort"]
            area2_sort = df.loc[i+1, "Sort"]
            if isNaN(area2):
                finalArea = area1
            elif not isNaN(area2) and area2_sort != 1:
                finalArea = area1
            elif not isNaN(area2) and area2_sort == 1:
                finalArea = area2
        if finalArea not in areaNames and finalArea != "":
            areaNames.append(finalArea)
            area_dict = {'Area ID'    : areaNames.index(finalArea),
                        'Area Name'   : finalArea
                        }
            area_list.append(area_dict)
    area_df = pd.DataFrame(area_list)
    area_df.fillna("", inplace=True)

    with pd.ExcelWriter(filepath, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        area_df.to_excel(writer, sheet_name='Machine Areas', index=False)
