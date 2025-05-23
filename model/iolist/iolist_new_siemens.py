import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo

# Create a new Excel workbook
# workbook = openpyxl.Workbook()

# # Select the default active sheet
# sheet = workbook.active

# # Define your column headers
# column_headers = [
#     ["Header 1", "Header 2", "Header 3"],
#     ["Subheader 1", "Subheader 2", "Subheader 3"],
# ]

# # Add the column headers to the sheet
# for row_index, headers in enumerate(column_headers, start=1):
#     for col_index, header_text in enumerate(headers, start=1):
#         sheet.cell(row=row_index, column=col_index, value=header_text)

# # Optionally, add some data below the headers
# data = [
#     [1, 2, 3],
#     [4, 5, 6],
#     [7, 8, 9],
# ]

# for row_index, row_data in enumerate(data, start=len(column_headers) + 1):
#     for col_index, value in enumerate(row_data, start=1):
#         sheet.cell(row=row_index, column=col_index, value=value)

# # Create a table from the data (optional)
# table = Table(displayName="MyTable", ref=sheet.dimensions)
# # style = TableStyleInfo(
# #     name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True
# # )
# # table.tableStyleInfo = style
# sheet.add_table(table)

# # Save the workbook
# workbook.save("excel_file_with_headers.xlsx")